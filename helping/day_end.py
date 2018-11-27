from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta

algos = ["SVM", "GradBoost", "AdaBoost", "LogReg", "MLP", "fixedSVM", "fixedMLP", "fixedGrad", "fixedAda", "fixedLog"]
orange = PatternFill(fill_type='solid', start_color='ED7D31', end_color='ED7D31')
blue = PatternFill(fill_type='solid', start_color='4472C4', end_color='4472C4')

def sum_up(letter, start, end):
    letter_sum = 0
    for i in range(start, end+1, 3):
        letter_sum += float(file[letter+str(i)].value)
    return letter_sum

for alg in algos:
    print(alg)
    file_name = '../testings/testing_' + alg + '.xlsx'
    file_book = load_workbook(file_name)
    file = file_book['Лист1']
    last_row = len(list(file.rows))
    counter = last_row - 6
    while file['B'+str(counter)].value==None:
        counter -= 3
        
    data = file['B'+str(counter)].value
    next_date = data + timedelta(days=1)
    
    bank_add=False
    if file['E'+str(counter)].value!=None and file['E'+str(counter)].value[:5]=="Bank=":
        bank_add=True

    file['A'+str(last_row+1)].value = 'RESULTS'
    file['A'+str(last_row+2)].number_format = 'dd.mm.yyyy'
    file['A'+str(last_row+2)].value = data
    file['A'+str(last_row+1)].fill = file['A'+str(last_row+2)].fill = orange

    file['B'+str(last_row+3)].number_format = 'dd.mm.yyyy'
    file['B'+str(last_row+3)].value = next_date
    file['C'+str(last_row+1)].value = file['C'+str(last_row+2)].value = '!!!!!!!!!!!!!'
    file['C'+str(last_row+1)].fill = file['C'+str(last_row+2)].fill = orange

    file['E'+str(last_row+1)].value = "=SUM(E"+str(counter+1)+":E"+str(last_row-2)+")"
    file['G'+str(last_row+1)].value = "=SUM(G"+str(counter+1)+":G"+str(last_row-2)+")"
    file['I'+str(last_row+1)].value = "=SUM(I"+str(counter+1)+":I"+str(last_row-2)+")"
    e_sum = sum_up('E', counter+1, last_row-2)
    g_sum = sum_up('G', counter+1, last_row-2)
    i_sum = sum_up('I', counter+1, last_row-2)
    if bank_add:
        file['E'+str(last_row+3)].value = "Bank="+str(float(file['E'+str(counter)].value[5:]) + e_sum)
        file['G'+str(last_row+3)].value = "Bank="+str(float(file['G'+str(counter)].value[5:]) + g_sum)
        file['I'+str(last_row+3)].value = "Bank="+str(float(file['I'+str(counter)].value[5:]) + i_sum)
    file['E'+str(last_row+1)].fill = file['G'+str(last_row+1)].fill = file['I'+str(last_row+1)].fill = orange
    file['E'+str(last_row+2)].fill = file['G'+str(last_row+2)].fill = file['I'+str(last_row+2)].fill = orange
    
    if alg!="SVM" and alg!="fixedSVM":
        file['K'+str(last_row+1)].value = "=SUM(K"+str(counter+1)+":K"+str(last_row-2)+")"
        if bank_add:
            k_sum = sum_up('K', counter+1, last_row-2)
            file['K'+str(last_row+3)].value = "Bank="+str(float(file['K'+str(counter)].value[5:]) + k_sum)
        file['K'+str(last_row+1)].fill = file['K'+str(last_row+2)].fill = orange

    for cell in file[last_row+3][:20]:
        cell.fill = blue

    file_book.save(file_name)
