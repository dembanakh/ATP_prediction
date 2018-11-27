from openpyxl import load_workbook

wb = load_workbook('../Data/2019_simple.xlsx')
ws = wb['2019_simple']

rankings = load_workbook('../Data/elo_rankings.xlsx')
elo = rankings['elo']

class Match(object):
    def __init__(self, atp=None, location=None, winner=None, loser=None, wrank=None, lrank=None, surface=None, date=None, cfw=None, cfl=None, round_n=None, series=None, \
                 wfsw=None, wssw=None, lfsw=None, lssw=None, wfrw=None, wsrw=None, lfrw=None, lsrw=None, wsrat=None, wrrat=None, lsrat=None, lrrat=None, \
                 waces=None, laces=None, wdf=None, ldf=None, wbp=None, lbp=None, welo=None, lelo=None, welosur=None, lelosur=None):
        self.atp = atp
        self.location = location
        self.winner = winner
        self.loser = loser
        self.surface = surface
        self.date = date
        self.round = round_n
        self.series = series
        self.welo = welo
        self.lelo = lelo
        self.welosur = welosur
        self.lelosur = lelosur
        

def adjust(series, round_n, atp):
    if series=='ATP250':
        s = 0.7
        if atp==52:
            if round_n=='1st Round':
                r = 0.75
            elif round_n=='2nd Round':
                r = 0.8
            elif round_n=='3rd Round':
                r = 0.8
            elif round_n=='Quarterfinals':
                r = 0.85
            elif round_n=='Semifinals':
                r = 0.9
            elif round_n=='The Final':
                r = 1
        else:
            if round_n=='1st Round':
                r = 0.8
            elif round_n=='2nd Round':
                r = 0.8
            elif round_n=='Quarterfinals':
                r = 0.85
            elif round_n=='Semifinals':
                r = 0.9
            elif round_n=='The Final':
                r = 1
    elif series=='ATP500':
        s = 0.75
        if atp in [24, 49]:
            if round_n=='1st Round':
                r = 0.75
            elif round_n=='2nd Round':
                r = 0.8
            elif round_n=='3rd Round':
                r = 0.8
            elif round_n=='Quarterfinals':
                r = 0.85
            elif round_n=='Semifinals':
                r = 0.9
            elif round_n=='The Final':
                r = 1
        else:
            if round_n=='1st Round':
                r = 0.8
            elif round_n=='2nd Round':
                r = 0.8
            elif round_n=='Quarterfinals':
                r = 0.85
            elif round_n=='Semifinals':
                r = 0.9
            elif round_n=='The Final':
                r = 1
    elif series=='Masters 1000':
        s = 0.85
        if atp in [19, 20]:
            if round_n=='1st Round':
                r = 0.75
            elif round_n=='2nd Round':
                r = 0.75
            elif round_n=='3rd Round':
                r = 0.8
            elif round_n=='4th Round':
                r = 0.8
            elif round_n=='Quarterfinals':
                r = 0.85
            elif round_n=='Semifinals':
                r = 0.9
            elif round_n=='The Final':
                r = 1
        else:
            if round_n=='1st Round':
                r = 0.75
            elif round_n=='2nd Round':
                r = 0.8
            elif round_n=='3rd Round':
                r = 0.8
            elif round_n=='Quarterfinals':
                r = 0.85
            elif round_n=='Semifinals':
                r = 0.9
            elif round_n=='The Final':
                r = 1
    elif series=='Grand Slam':
        s = 1
        if round_n=='1st Round':
            r = 0.75
        elif round_n=='2nd Round':
            r = 0.75
        elif round_n=='3rd Round':
            r = 0.8
        elif round_n=='4th Round':
            r = 0.8
        elif round_n=='Quarterfinals':
            r = 0.85
        elif round_n=='Semifinals':
            r = 0.9
        elif round_n=='The Final':
            r = 1
    elif series=='Masters Cup':
        s = 0.9
        if round_n=='Round Robin':
            r = 0.85
        elif round_n=='Semifinals':
            r = 0.9
        elif round_n=='The Final':
            r = 1

    return r*s


elo_surface = {'Clay': "../Data/elo_rankings_1.xlsx",
               'Hard': "../Data/elo_rankings_2.xlsx",
               'Grass': "../Data/elo_rankings_3.xlsx"} 

players = {}
for row in list(elo.rows):
        if row[0].value==None:
            break
        players[row[0].value] = int(row[1].value)

start, end = 0, 0

surface = load_workbook(elo_surface[list(ws.rows)[start][6].value])
elo_sur = surface['Лист1']

players_sur = {}
for row in list(elo_sur.rows):
    if row[0].value==None:
            break
    players_sur[row[0].value] = int(row[1].value)

    
calendar = []
i_s = [i for i in range(start, end)]
for row in list(ws.rows)[start:end]:
    args = [cell.value for cell in row][:30]
    if row[2].value not in players:
        players[row[2].value] = 1500
    if row[2].value not in players_sur:
        players_sur[row[2].value] = 1500
    row[30].value = players[row[2].value]
    row[34].value = players_sur[row[2].value]
    if row[3].value not in players:
        players[row[3].value] = 1500
    if row[3].value not in players_sur:
        players_sur[row[3].value] = 1500
    row[31].value = players[row[3].value]
    row[35].value = players_sur[row[3].value]
    args += [row[30].value, row[31].value, row[34].value, row[35].value]
    match = Match(*args)
    calendar.append(match)

for i, match in enumerate(calendar):
    e1 = 1/(1 + 10**((match.lelo-match.welo)/400))
    e2 = 1-e1

    e1sur = 1/(1 + 10**((match.lelosur-match.welosur)/400))
    e2sur = 1-e1sur

    K = 32*adjust(match.series, match.round, match.atp)

    wdelta = round(K*e2, 0)
    ldelta = round(K*(-e2), 0)

    wdeltasur = round(K*e2sur, 0)
    ldeltasur = round(K*(-e2sur), 0)

    ws['AG'+str(i_s[i]+1)] = wdelta
    ws['AH'+str(i_s[i]+1)] = ldelta

    ws['AK'+str(i_s[i]+1)] = wdeltasur
    ws['AL'+str(i_s[i]+1)] = ldeltasur

    players[match.winner] = match.welo+wdelta
    players[match.loser] = match.lelo+ldelta

    players_sur[match.winner] = match.welosur+wdeltasur
    players_sur[match.loser] = match.lelosur+ldeltasur

    print(match.winner, match.loser)


for i, player in enumerate(players):
    elo['A'+str(i+1)] = player
    elo['B'+str(i+1)] = players[player]

for i, player in enumerate(players_sur):
    elo_sur['A'+str(i+1)] = player
    elo_sur['B'+str(i+1)] = players_sur[player]

wb.save('../Data/2019_simple.xlsx')
rankings.save('../Data/elo_rankings.xlsx')
surface.save(elo_surface[list(ws.rows)[start][6].value])
