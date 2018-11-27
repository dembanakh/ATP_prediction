from openpyxl import load_workbook
import bs4
import requests
import datetime

page = requests.get("https://www.atpworldtour.com/en/tournaments").text
content = bs4.BeautifulSoup(page, "html.parser")

wb = load_workbook('../Data/2019_simple.xlsx')
ws = wb['2019_simple']

month_tour = content.select('div.content-accordion')
urls = []

for tourney in content.select('tr.tourney-result'):
    name = tourney.select('.title-content')[0].find(class_='tourney-location').get_text().split(', ')[0][14:32]
    if name=='Multiple Locations' or name=='Milan':
        continue
    url = tourney.select('.title-content')[0].select('a')[0].get('href')[15:-8]
    urls.append((name, url))

for i in range(0, 3):
    print("Tournament no. {} - {}...".format(i+1, urls[i][0]))
    res_page = "https://www.atpworldtour.com/en/scores/current" + urls[i][1] + "results"
    res_text = requests.get(res_page).text
    res = bs4.BeautifulSoup(res_text, "html.parser")
    try:
        rounds = res.select('table.day-table')[0]
    except IndexError:
        print("no rounds")
        continue
    atp_start = 2
    if len(rounds.find_all('thead'))<=9:
        start = len(rounds.find_all('thead'))-3
    elif len(rounds.find_all('thead'))==10:
        start = 6
    for g in range(atp_start, len(ws['A'])):
        if ws['B'+str(g)].value==urls[i][0]:
            atp_start = g
            break
        
    for j in range(start, -1, -1):
        print(" Round no.{}...".format(start+1-j))
        block = rounds.find_all('tbody')[j]
        for match in block.find_all('tr'):
            player1 = ' '.join(match.select('td.day-table-name')[0].get_text()[1:-1].split(' ')[1:]).lower()
            player2 = ' '.join(match.select('td.day-table-name')[1].get_text()[1:-1].split(' ')[1:]).lower()
            link = match.select('td.day-table-score')[0].find('a').get('href')
            '''scorebox = match.find(id="completedScoreBox")
            #print(scorebox.select('tr')[0].select('td')[0]["class"])
            if scorebox.select('tr')[0].select('td')[0]["class"][0]=="won-game":
                who = 1
                player1 = match.select('.player-left-name')[0].select('.last-name')[0].get_text().strip().lower()
                player2 = match.select('.player-right-name')[0].select('.last-name')[0].get_text().strip().lower()
            elif scorebox.select('tr')[1].select('td')[0]["class"][0]=="won-game":
                who = 2
                player2 = match.select('.player-left-name')[0].select('.last-name')[0].get_text().strip().lower()
                player1 = match.select('.player-right-name')[0].select('.last-name')[0].get_text().strip().lower()'''
            for g in range(atp_start, len(ws['A'])+1):
                if player1==' '.join(ws['C'+str(g)].value.split(' ')[:-1]).lower() and player2==' '.join(ws['D'+str(g)].value.split(' ')[:-1]).lower() and \
                   ws['B'+str(g)].value==urls[i][0]:
                    if ws['M'+str(g)].value!=None:
                        break
                    if link!=None:
                         match_page = "https://www.atpworldtour.com"+link
                    else:
                        break
                    print("   {} - {}".format(player1, player2))
                    match_text = requests.get(match_page).text
                    match_ = bs4.BeautifulSoup(match_text, "html.parser")
                    stats = match_.find(id="completedMatchStats")
                    scorebox = match_.find(id="completedScoreBox")
                    if scorebox.select('tr')[0].select('td')[0]["class"]=="won-game":
                        who = 1
                    elif scorebox.select('tr')[1].select('td')[0]["class"]=="won-game":
                        who = 2
                    ### FIRST SERVE POINTS WON
                    first_serve_won = stats.select('tr')[5]
                    if who==1:
                        fswon_1 = first_serve_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        fswon_2 = first_serve_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    else:
                        fswon_2 = first_serve_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        fswon_1 = first_serve_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    ws['M'+str(g)] = fswon_1
                    ws['O'+str(g)] = fswon_2
                    ### SECOND SERVE POINTS WON
                    second_serve_won = stats.select('tr')[6]
                    if who==1:
                        sswon_1 = second_serve_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        sswon_2 = second_serve_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    else:
                        sswon_2 = second_serve_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        sswon_1 = second_serve_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    ws['N'+str(g)] = sswon_1
                    ws['P'+str(g)] = sswon_2
                    ### FIRST RETURN POINTS WON
                    first_return_won = stats.select('tr')[11]
                    if who==1:
                        frwon_1 = first_return_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        frwon_2 = first_return_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    else:
                        frwon_2 = first_return_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        frwon_1 = first_return_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    ws['Q'+str(g)] = frwon_1
                    ws['S'+str(g)] = frwon_2
                    ### SECOND RETURN POINTS WON
                    second_return_won = stats.select('tr')[12]
                    if who==1:
                        srwon_1 = second_return_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        srwon_2 = second_return_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    else:
                        srwon_2 = second_return_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        srwon_1 = second_return_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    ws['R'+str(g)] = srwon_1
                    ws['T'+str(g)] = srwon_2
                    ### SERVE RATING
                    serve_rating = stats.select('tr')[1]
                    sr_1 = serve_rating.select('td.match-stats-number-left')[0].find('span').get_text().strip()
                    sr_2 = serve_rating.select('td.match-stats-number-right')[0].find('span').get_text().strip()
                    ws['U'+str(g)] = sr_1
                    ws['W'+str(g)] = sr_2
                    ### RETURN RATING
                    return_rating = stats.select('tr')[10]
                    if who==1:
                        rr_1 = return_rating.select('td.match-stats-number-left')[0].find('span').get_text().strip()
                        rr_2 = return_rating.select('td.match-stats-number-right')[0].find('span').get_text().strip()
                    else:
                        rr_2 = return_rating.select('td.match-stats-number-left')[0].find('span').get_text().strip()
                        rr_1 = return_rating.select('td.match-stats-number-right')[0].find('span').get_text().strip()
                    ws['V'+str(g)] = rr_1
                    ws['X'+str(g)] = rr_2
                    ### ACES Per Service Game
                    aces = stats.select('tr')[2]
                    srv = stats.select('tr')[8]
                    if who==1:
                        aces_1 = int(aces.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                        aces_2 = int(aces.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                        srv_1 = int(srv.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                        srv_2 = int(srv.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                    else:
                        aces_2 = int(aces.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                        aces_1 = int(aces.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                        srv_2 = int(srv.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                        srv_1 = int(srv.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                    ws['Y'+str(g)] = round(aces_1/srv_1, 3)
                    ws['Z'+str(g)] = round(aces_2/srv_2, 3)
                    ### DOUBLE FAULTS Per Service Game
                    df = stats.select('tr')[3]
                    if who==1:
                        df_1 = int(df.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                        df_2 = int(df.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                    else:
                        df_2 = int(df.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                        df_1 = int(df.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                    ws['AA'+str(g)] = round(df_1/srv_1, 3)
                    ws['AB'+str(g)] = round(df_2/srv_2, 3)
                    ### BREAK POINTS Converted
                    bp_conv = stats.select('tr')[13]
                    if who==1:
                        bp_1 = bp_conv.select('td.match-stats-number-left')[0].find('span').get_text().strip()
                        bp_2 = bp_conv.select('td.match-stats-number-right')[0].find('span').get_text().strip()
                    else:
                        bp_2 = bp_conv.select('td.match-stats-number-left')[0].find('span').get_text().strip()
                        bp_1 = bp_conv.select('td.match-stats-number-right')[0].find('span').get_text().strip()
                    ws['AC'+str(g)] = bp_1
                    ws['AD'+str(g)] = bp_2
                    break
                elif (ws['B'+str(g)].value==urls[i][0] and player2==' '.join(ws['D'+str(g)].value.split(' ')[:-1]).lower() and ' '.join(player1.split()[1:]).lower()==' '.join(ws['C'+str(g)].value.split()[:-1]).lower()) or (ws['B'+str(g)].value==urls[i][0] and player1==' '.join(ws['C'+str(g)].value.split(' ')[:-1]).lower() and ' '.join(player2.split()[1:]).lower()==' '.join(ws['D'+str(g)].value.split()[:-1]).lower()):
                    if ws['M'+str(g)].value!=None:
                        break
                    if link!=None:
                         match_page = "https://www.atpworldtour.com"+link
                    else:
                        break
                    print("   {} - {}".format(player1, player2))
                    match_text = requests.get(match_page).text
                    match_ = bs4.BeautifulSoup(match_text, "html.parser")
                    stats = match_.find(id="completedMatchStats")
                    scorebox = match_.find(id="completedScoreBox")
                    if scorebox.select('tr')[0].select('td')[0]["class"]=="won-game":
                        who = 1
                    elif scorebox.select('tr')[1].select('td')[0]["class"]=="won-game":
                        who = 2
                    ### FIRST SERVE POINTS WON
                    first_serve_won = stats.select('tr')[5]
                    if who==1:
                        fswon_1 = first_serve_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        fswon_2 = first_serve_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    else:
                        fswon_2 = first_serve_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        fswon_1 = first_serve_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    ws['M'+str(g)] = fswon_1
                    ws['O'+str(g)] = fswon_2
                    ### SECOND SERVE POINTS WON
                    second_serve_won = stats.select('tr')[6]
                    if who==1:
                        sswon_1 = second_serve_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        sswon_2 = second_serve_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    else:
                        sswon_2 = second_serve_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        sswon_1 = second_serve_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    ws['N'+str(g)] = sswon_1
                    ws['P'+str(g)] = sswon_2
                    ### FIRST RETURN POINTS WON
                    first_return_won = stats.select('tr')[11]
                    if who==1:
                        frwon_1 = first_return_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        frwon_2 = first_return_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    else:
                        frwon_2 = first_return_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        frwon_1 = first_return_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    ws['Q'+str(g)] = frwon_1
                    ws['S'+str(g)] = frwon_2
                    ### SECOND RETURN POINTS WON
                    second_return_won = stats.select('tr')[12]
                    if who==1:
                        srwon_1 = second_return_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        srwon_2 = second_return_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    else:
                        srwon_2 = second_return_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        srwon_1 = second_return_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    ws['R'+str(g)] = srwon_1
                    ws['T'+str(g)] = srwon_2
                    ### SERVE RATING
                    serve_rating = stats.select('tr')[1]
                    sr_1 = serve_rating.select('td.match-stats-number-left')[0].find('span').get_text().strip()
                    sr_2 = serve_rating.select('td.match-stats-number-right')[0].find('span').get_text().strip()
                    ws['U'+str(g)] = sr_1
                    ws['W'+str(g)] = sr_2
                    ### RETURN RATING
                    return_rating = stats.select('tr')[10]
                    if who==1:
                        rr_1 = return_rating.select('td.match-stats-number-left')[0].find('span').get_text().strip()
                        rr_2 = return_rating.select('td.match-stats-number-right')[0].find('span').get_text().strip()
                    else:
                        rr_2 = return_rating.select('td.match-stats-number-left')[0].find('span').get_text().strip()
                        rr_1 = return_rating.select('td.match-stats-number-right')[0].find('span').get_text().strip()
                    ws['V'+str(g)] = rr_1
                    ws['X'+str(g)] = rr_2
                    ### ACES Per Service Game
                    aces = stats.select('tr')[2]
                    srv = stats.select('tr')[8]
                    if who==1:
                        aces_1 = int(aces.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                        aces_2 = int(aces.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                        srv_1 = int(srv.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                        srv_2 = int(srv.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                    else:
                        aces_2 = int(aces.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                        aces_1 = int(aces.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                        srv_2 = int(srv.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                        srv_1 = int(srv.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                    ws['Y'+str(g)] = round(aces_1/srv_1, 3)
                    ws['Z'+str(g)] = round(aces_2/srv_2, 3)
                    ### DOUBLE FAULTS Per Service Game
                    df = stats.select('tr')[3]
                    if who==1:
                        df_1 = int(df.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                        df_2 = int(df.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                    else:
                        df_2 = int(df.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                        df_1 = int(df.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                    ws['AA'+str(g)] = round(df_1/srv_1, 3)
                    ws['AB'+str(g)] = round(df_2/srv_2, 3)
                    ### BREAK POINTS Converted
                    bp_conv = stats.select('tr')[13]
                    if who==1:
                        bp_1 = bp_conv.select('td.match-stats-number-left')[0].find('span').get_text().strip()
                        bp_2 = bp_conv.select('td.match-stats-number-right')[0].find('span').get_text().strip()
                    else:
                        bp_2 = bp_conv.select('td.match-stats-number-left')[0].find('span').get_text().strip()
                        bp_1 = bp_conv.select('td.match-stats-number-right')[0].find('span').get_text().strip()
                    ws['AC'+str(g)] = bp_1
                    ws['AD'+str(g)] = bp_2
                    break
                                  
    wb.save('../Data/2019_simple.xlsx')

'''print("Tournament no. {} - {}...".format(67, "ATP Finals"))
for i in range(1, 2):
    url = "https://www.atpworldtour.com/en/scores/2018/605/MS00" + str(i) + "/match-stats"
    res_text = requests.get(url).text
    res = bs4.BeautifulSoup(res_text, "html.parser")
    try:
        match = res.select('.modal-scores-match-stats-table')[0]
    except:
        continue
    scorebox = match.find(id="completedScoreBox")
    #print(scorebox.select('tr')[0].select('td')[0]["class"])
    if scorebox.select('tr')[0].select('td')[0]["class"][0]=="won-game":
        who = 1
        player1 = match.select('.player-left-name')[0].select('.last-name')[0].get_text().strip().lower()
        player2 = match.select('.player-right-name')[0].select('.last-name')[0].get_text().strip().lower()
    elif scorebox.select('tr')[1].select('td')[0]["class"][0]=="won-game":
        who = 2
        player2 = match.select('.player-left-name')[0].select('.last-name')[0].get_text().strip().lower()
        player1 = match.select('.player-right-name')[0].select('.last-name')[0].get_text().strip().lower()
    for g in range(2616, 2620):
                if player1==' '.join(ws['C'+str(g)].value.split(' ')[:-1]).lower() and player2==' '.join(ws['D'+str(g)].value.split(' ')[:-1]).lower():
                    if ws['AC'+str(g)].value!=ws['AD'+str(g)].value:
                        break
                    print("   {} - {}".format(player1, player2))
                    stats = match.find(id="completedMatchStats")
                    ### FIRST SERVE POINTS WON
                    first_serve_won = stats.select('tr')[5]
                    if who==1:
                        fswon_1 = first_serve_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        fswon_2 = first_serve_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    else:
                        fswon_2 = first_serve_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        fswon_1 = first_serve_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    ws['M'+str(g)] = fswon_1
                    ws['O'+str(g)] = fswon_2
                    ### SECOND SERVE POINTS WON
                    second_serve_won = stats.select('tr')[6]
                    if who==1:
                        sswon_1 = second_serve_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        sswon_2 = second_serve_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    else:
                        sswon_2 = second_serve_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        sswon_1 = second_serve_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    ws['N'+str(g)] = sswon_1
                    ws['P'+str(g)] = sswon_2
                    ### FIRST RETURN POINTS WON
                    first_return_won = stats.select('tr')[11]
                    if who==1:
                        frwon_1 = first_return_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        frwon_2 = first_return_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    else:
                        frwon_2 = first_return_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        frwon_1 = first_return_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    ws['Q'+str(g)] = frwon_1
                    ws['S'+str(g)] = frwon_2
                    ### SECOND RETURN POINTS WON
                    second_return_won = stats.select('tr')[12]
                    if who==1:
                        srwon_1 = second_return_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        srwon_2 = second_return_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    else:
                        srwon_2 = second_return_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                        srwon_1 = second_return_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    ws['R'+str(g)] = srwon_1
                    ws['T'+str(g)] = srwon_2
                    ### SERVE RATING
                    serve_rating = stats.select('tr')[1]
                    sr_1 = serve_rating.select('td.match-stats-number-left')[0].find('span').get_text().strip()
                    sr_2 = serve_rating.select('td.match-stats-number-right')[0].find('span').get_text().strip()
                    ws['U'+str(g)] = sr_1
                    ws['W'+str(g)] = sr_2
                    ### RETURN RATING
                    return_rating = stats.select('tr')[10]
                    if who==1:
                        rr_1 = return_rating.select('td.match-stats-number-left')[0].find('span').get_text().strip()
                        rr_2 = return_rating.select('td.match-stats-number-right')[0].find('span').get_text().strip()
                    else:
                        rr_2 = return_rating.select('td.match-stats-number-left')[0].find('span').get_text().strip()
                        rr_1 = return_rating.select('td.match-stats-number-right')[0].find('span').get_text().strip()
                    ws['V'+str(g)] = rr_1
                    ws['X'+str(g)] = rr_2
                    ### ACES Per Service Game
                    aces = stats.select('tr')[2]
                    srv = stats.select('tr')[8]
                    if who==1:
                        aces_1 = int(aces.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                        aces_2 = int(aces.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                        srv_1 = int(srv.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                        srv_2 = int(srv.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                    else:
                        aces_2 = int(aces.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                        aces_1 = int(aces.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                        srv_2 = int(srv.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                        srv_1 = int(srv.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                    ws['Y'+str(g)] = round(aces_1/srv_1, 3)
                    ws['Z'+str(g)] = round(aces_2/srv_2, 3)
                    ### DOUBLE FAULTS Per Service Game
                    df = stats.select('tr')[3]
                    if who==1:
                        df_1 = int(df.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                        df_2 = int(df.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                    else:
                        df_2 = int(df.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                        df_1 = int(df.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                    ws['AA'+str(g)] = round(df_1/srv_1, 3)
                    ws['AB'+str(g)] = round(df_2/srv_2, 3)
                    ### BREAK POINTS Converted
                    bp_conv = stats.select('tr')[13]
                    if who==1:
                        bp_1 = bp_conv.select('td.match-stats-number-left')[0].find('span').get_text().strip()
                        bp_2 = bp_conv.select('td.match-stats-number-right')[0].find('span').get_text().strip()
                    else:
                        bp_2 = bp_conv.select('td.match-stats-number-left')[0].find('span').get_text().strip()
                        bp_1 = bp_conv.select('td.match-stats-number-right')[0].find('span').get_text().strip()
                    ws['AC'+str(g)] = bp_1
                    ws['AD'+str(g)] = bp_2
                    break
                elif (player2==' '.join(ws['D'+str(g)].value.split(' ')[:-1]).lower() and ' '.join(player1.split()[1:]).lower()==' '.join(ws['C'+str(g)].value.split()[:-1]).lower()) or (player1==' '.join(ws['C'+str(g)].value.split(' ')[:-1]).lower() and ' '.join(player2.split()[1:]).lower()==' '.join(ws['D'+str(g)].value.split()[:-1]).lower()):
                    if ws['AC'+str(g)].value!=ws['AD'+str(g)].value:
                        break
                    print("   {} - {}".format(player1, player2))
                    stats = match.find(id="completedMatchStats")
                    ### FIRST SERVE POINTS WON
                    first_serve_won = stats.select('tr')[5]
                    fswon_1 = first_serve_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                    fswon_2 = first_serve_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    ws['M'+str(g)] = fswon_1
                    ws['O'+str(g)] = fswon_2
                    ### SECOND SERVE POINTS WON
                    second_serve_won = stats.select('tr')[6]
                    sswon_1 = second_serve_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                    sswon_2 = second_serve_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    ws['N'+str(g)] = sswon_1
                    ws['P'+str(g)] = sswon_2
                    ### FIRST RETURN POINTS WON
                    first_return_won = stats.select('tr')[11]
                    frwon_1 = first_return_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                    frwon_2 = first_return_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    ws['Q'+str(g)] = frwon_1
                    ws['S'+str(g)] = frwon_2
                    ### SECOND RETURN POINTS WON
                    second_return_won = stats.select('tr')[12]
                    srwon_1 = second_return_won.select('td.match-stats-number-left')[0].find('span').get_text()[22:-18]
                    srwon_2 = second_return_won.select('td.match-stats-number-right')[0].find('span').get_text()[22:-18]
                    ws['R'+str(g)] = srwon_1
                    ws['T'+str(g)] = srwon_2
                    ### SERVE RATING
                    serve_rating = stats.select('tr')[1]
                    sr_1 = serve_rating.select('td.match-stats-number-left')[0].find('span').get_text().strip()
                    sr_2 = serve_rating.select('td.match-stats-number-right')[0].find('span').get_text().strip()
                    ws['U'+str(g)] = sr_1
                    ws['W'+str(g)] = sr_2
                    ### RETURN RATING
                    return_rating = stats.select('tr')[10]
                    rr_1 = return_rating.select('td.match-stats-number-left')[0].find('span').get_text().strip()
                    rr_2 = return_rating.select('td.match-stats-number-right')[0].find('span').get_text().strip()
                    ws['V'+str(g)] = rr_1
                    ws['X'+str(g)] = rr_2
                    ### ACES Per Service Game
                    aces = stats.select('tr')[2]
                    srv = stats.select('tr')[8]
                    aces_1 = int(aces.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                    aces_2 = int(aces.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                    srv_1 = int(srv.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                    srv_2 = int(srv.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                    ws['Y'+str(g)] = round(aces_1/srv_1, 3)
                    ws['Z'+str(g)] = round(aces_2/srv_2, 3)
                    ### DOUBLE FAULTS Per Service Game
                    df = stats.select('tr')[3]
                    df_1 = int(df.select('td.match-stats-number-left')[0].find('span').get_text().strip())
                    df_2 = int(df.select('td.match-stats-number-right')[0].find('span').get_text().strip())
                    ws['AA'+str(g)] = round(df_1/srv_1, 3)
                    ws['AB'+str(g)] = round(df_2/srv_2, 3)
                    ### BREAK POINTS Converted
                    bp_conv = stats.select('tr')[13]
                    bp_1 = bp_conv.select('td.match-stats-number-left')[0].find('span').get_text().strip()
                    bp_2 = bp_conv.select('td.match-stats-number-right')[0].find('span').get_text().strip()
                    ws['AC'+str(g)] = bp_1
                    ws['AD'+str(g)] = bp_2
                    break
                                  
    wb.save('../Data/2018_simple.xlsx')'''
