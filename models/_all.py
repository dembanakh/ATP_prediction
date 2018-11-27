import numpy as np
import sys
from sklearn.svm import SVC
from sklearn.ensemble import GradientBoostingClassifier as GBC
from sklearn.ensemble import AdaBoostClassifier as ABC
from sklearn.linear_model import LogisticRegression as Log
from sklearn.neural_network import MLPClassifier as MLP
from sklearn.metrics import make_scorer
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import MinMaxScaler
from sklearn.preprocessing import StandardScaler
from datetime import datetime, timedelta
import matplotlib.pyplot as plt

if sys.argv[-1]=='_all.py':
    print("\nType in an index of the starting match!\n")
    raise 

import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)

wb2017 = load_workbook('../Data/2017_simple.xlsx')
ws2017 = wb2017['2017_simple']

wb2018 = load_workbook('../Data/2018_simple.xlsx')
ws2018 = wb2018['2018_simple']

wb2019 = load_workbook('../Data/2019_simple.xlsx')
ws2019 = wb2019['2019_simple']

blue = PatternFill(fill_type='solid', start_color='4472C4', end_color='4472C4')

f = 0.8

def not_NA_rank(rank):
    if rank == 'N/A':
        return 2000
    else:
        return rank

def write_profits(algo, players, coeffs, profits):
    file_name = '../testings/testing_' + algo +'.xlsx'
    testing_book = load_workbook(file_name)
    testing = testing_book['Лист1']
    last_row = len(list(testing.rows))
    
    testing['A'+str(last_row+1)].value = players[0].split()[0]
    testing['A'+str(last_row+2)].value = players[1].split()[0]
    testing['C'+str(last_row+1)].value = coeffs[0]
    testing['C'+str(last_row+2)].value = coeffs[1]

    testing['E'+str(last_row+1)].value = profits[0]
    testing['G'+str(last_row+1)].value = profits[1]
    testing['I'+str(last_row+1)].value = profits[2]
    if len(profits)==4:
        testing['K'+str(last_row+1)].value = profits[3]

    for cell in testing[last_row+3][:20]:
        cell.fill = blue

    testing_book.save(file_name)
        
    

class Match(object):
    def __init__(self, atp=None, location=None, winner=None, loser=None, wrank=None, lrank=None, surface=None, date=None, cfw=None, cfl=None, round_n=None, series=None, \
                 wfsw=None, wssw=None, lfsw=None, lssw=None, wfrw=None, wsrw=None, lfrw=None, lsrw=None, wsrat=None, wrrat=None, lsrat=None, lrrat=None, \
                 waces=None, laces=None, wdf=None, ldf=None, wbp=None, lbp=None, welo=None, lelo=None, bug1=None, bug2=None, welosur=None, lelosur=None):
        self.atp = atp
        self.location = location
        self.winner = winner
        self.loser = loser
        self.wrank = not_NA_rank(wrank)
        self.lrank = not_NA_rank(lrank)
        self.surface = surface
        self.date = date
        self.cfw = cfw
        self.cfl = cfl
        self.round = round_n
        self.series = series
        self.wfsw=wfsw
        self.wssw=wssw
        self.lfsw=lfsw
        self.lssw=lssw
        self.wfrw=wfrw
        self.wsrw=wsrw
        self.lfrw=lfrw
        self.lsrw=lsrw
        self.waces=waces
        self.laces=laces
        self.wdf=wdf
        self.ldf=ldf
        self.wbp=wbp
        self.lbp=lbp
        self.welo=welo
        self.lelo=lelo
        self.welosur=welosur
        self.lelosur=lelosur
    def log(self):
        return dict(atp=self.atp, location=self.location, winner=self.winner, loser=self.loser, wrank=self.wrank, lrank=self.lrank, surface=self.surface, date=self.date, round_n=self.round, series=self.series)

surf_into_num = {'Clay': 0, 'Hard': 0.5, 'Grass': 1}

calendar = []
for row in list(ws2017.rows)[1:]:
    args = [cell.value for cell in row][:36]
    match = Match(*args)
    calendar.append(match)
for row in list(ws2018.rows)[1:]:
    args = [cell.value for cell in row][:36]
    match = Match(*args)
    calendar.append(match)
for row in list(ws2019.rows)[1:]:
    args = [cell.value for cell in row][:36]
    match = Match(*args)
    calendar.append(match)

def wins_per_surface(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    surface = match.surface
    i = 1
    wins = 0
    losses = 0
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match_temp.surface != match.surface:
            i += 1
            continue
        delta = match.date - match_temp.date
        if match.surface == 'Grass':
            t = delta.days/280
        else:
            t = delta.days/7
        if match_temp.winner == player:
            wins += min(f**t, f)
        elif match_temp.loser == player:
            losses += min(f**t, f)
        i += 1
        counter_matches -= 1
    if wins==0:
        return 0
    else:
        return round(wins/(wins+losses), 3)

def wins_percent(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    wins = 0
    losses = 0
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        delta = match.date - match_temp.date
        t = delta.days/7
        if match_temp.winner == player:
            wins += min(f**t, f)
        elif match_temp.loser == player:
            losses += min(f**t, f)
        i += 1
        counter_matches -= 1
    if wins==0:
        return 0
    else:
        return round(wins/(wins+losses), 3)

def av_first_serve(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    first_serve = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        delta = match.date - match_temp.date
        t = delta.days/28
        if match_temp.winner==player:
            first_serve.append((int(match_temp.wfsw[:-1])/100)*min(f**t, f))
        else:
            first_serve.append((int(match_temp.lfsw[:-1])/100)*min(f**t, f))
        i += 1
        counter_matches -= 1
    if len(first_serve)>0:
        return round(sum(first_serve)/len(first_serve), 3)
    else:
        return 0

def av_second_serve(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    second_serve = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        delta = match.date - match_temp.date
        t = delta.days/28
        if match_temp.winner==player:
            second_serve.append((int(match_temp.wssw[:-1])/100)*min(f**t, f))
        else:
            second_serve.append((int(match_temp.lssw[:-1])/100)*min(f**t, f))
        i += 1
        counter_matches -= 1
    if len(second_serve)>0:
        return round(sum(second_serve)/len(second_serve), 3)
    else:
        return 0

def av_first_serve_surface(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    first_serve = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match.surface != match_temp.surface:
            i += 1
            continue
        delta = match.date - match_temp.date
        if match.surface == 'Grass':
            t = 0
        else:
            t = delta.days/28
        if match_temp.winner==player:
            first_serve.append((int(match_temp.wfsw[:-1])/100)*min(f**t, f))
        else:
            first_serve.append((int(match_temp.lfsw[:-1])/100)*min(f**t, f))
        i += 1
        counter_matches -= 1
    if len(first_serve)>0:
        return round(sum(first_serve)/len(first_serve), 3)
    else:
        return 0

def av_second_serve_surface(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    second_serve = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match.surface != match_temp.surface:
            i += 1
            continue
        delta = match.date - match_temp.date
        if match.surface == 'Grass':
            t = 0
        else:
            t = delta.days/28
        if match_temp.winner==player:
            second_serve.append((int(match_temp.wssw[:-1])/100)*min(f**t, f))
        else:
            second_serve.append((int(match_temp.lssw[:-1])/100)*min(f**t, f))
        i += 1
        counter_matches -= 1
    if len(second_serve)>0:
        return round(sum(second_serve)/len(second_serve), 3)
    else:
        return 0

def av_first_return(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    first_serve = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        delta = match.date - match_temp.date
        t = delta.days/28
        if match_temp.winner==player:
            first_serve.append((int(match_temp.wfrw[:-1])/100)*min(f**t, f))
        else:
            first_serve.append((int(match_temp.lfrw[:-1])/100)*min(f**t, f))
        i += 1
        counter_matches -= 1
    if len(first_serve)>0:
        return round(sum(first_serve)/len(first_serve), 3)
    else:
        return 0

def av_second_return(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    second_serve = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        delta = match.date - match_temp.date
        t = delta.days/28
        if match_temp.winner==player:
            second_serve.append((int(match_temp.wsrw[:-1])/100)*min(f**t, f))
        else:
            second_serve.append((int(match_temp.lsrw[:-1])/100)*min(f**t, f))
        i += 1
        counter_matches -= 1
    if len(second_serve)>0:
        return round(sum(second_serve)/len(second_serve), 3)
    else:
        return 0

def av_first_return_surface(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    first_serve = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match.surface != match_temp.surface:
            i += 1
            continue
        delta = match.date - match_temp.date
        if match.surface == 'Grass':
            t = 0
        else:
            t = delta.days/28
        if match_temp.winner==player:
            first_serve.append((int(match_temp.wfrw[:-1])/100)*min(f**t, f))
        else:
            first_serve.append((int(match_temp.lfrw[:-1])/100)*min(f**t, f))
        i += 1
        counter_matches -= 1
    if len(first_serve)>0:
        return round(sum(first_serve)/len(first_serve), 3)
    else:
        return 0

def av_second_return_surface(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    second_serve = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match.surface != match_temp.surface:
            i += 1
            continue
        delta = match.date - match_temp.date
        if match.surface == 'Grass':
            t = 0
        else:
            t = delta.days/28
        if match_temp.winner==player:
            second_serve.append((int(match_temp.wsrw[:-1])/100)*min(f**t, f))
        else:
            second_serve.append((int(match_temp.lsrw[:-1])/100)*min(f**t, f))
        i += 1
        counter_matches -= 1
    if len(second_serve)>0:
        return round(sum(second_serve)/len(second_serve), 3)
    else:
        return 0

def av_aces(match, player):
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    aces = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match_temp.winner==player:
            aces.append(float(match_temp.waces))
        else:
            aces.append(float(match_temp.laces))
        i += 1
        counter_matches -= 1
    if len(aces)>0:
        return round(sum(aces)/len(aces), 3)
    else:
        return 0

def av_aces_surface(match, player):
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    aces = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match.surface != match_temp.surface:
            i += 1
            continue
        if match_temp.winner==player:
            aces.append(float(match_temp.waces))
        else:
            aces.append(float(match_temp.laces))
        i += 1
        counter_matches -= 1
    if len(aces)>0:
        return round(sum(aces)/len(aces), 3)
    else:
        return 0

def av_dfs(match, player):
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    dfs = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match_temp.winner==player:
            dfs.append(float(match_temp.wdf))
        else:
            dfs.append(float(match_temp.ldf))
        i += 1
        counter_matches -= 1
    if len(dfs)>0:
        return round(sum(dfs)/len(dfs), 5)
    else:
        return 0

def av_dfs_surface(match, player):
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    dfs = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match.surface != match_temp.surface:
            i += 1
            continue
        if match_temp.winner==player:
            dfs.append(float(match_temp.wdf))
        else:
            dfs.append(float(match_temp.ldf))
        i += 1
        counter_matches -= 1
    if len(dfs)>0:
        return round(sum(dfs)/len(dfs), 5)
    else:
        return 0

def av_bps(match, player):
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    bps = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match_temp.winner==player:
            bps.append(int(match_temp.wbp[:-1])/100)
        else:
            bps.append(int(match_temp.lbp[:-1])/100)
        i += 1
        counter_matches -= 1
    if len(bps)>0:
        return round(sum(bps)/len(bps), 3)
    else:
        return 0

def av_bps_surface(match, player):
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    bps = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match.surface != match_temp.surface:
            i += 1
            continue
        if match_temp.winner==player:
            bps.append(int(match_temp.wbp[:-1])/100)
        else:
            bps.append(int(match_temp.lbp[:-1])/100)
        i += 1
        counter_matches -= 1
    if len(bps)>0:
        return round(sum(bps)/len(bps), 3)
    else:
        return 0
    
algs1 = {
        "svm": 1000,
        "grad": 1000,
        "ada": 1000,
        "log": 1000,
        "mlp": 1000,
        "fixedsvm": 1000,
        "fixedmlp": 1000,
        "fixedgrad": 1000,
        "fixedada": 1000,
        "fixedlog": 1000}
algs2 = {
        "svm": 1000,
        "grad": 1000,
        "ada": 1000,
        "log": 1000,
        "mlp": 1000,
        "fixedsvm": 1000,
        "fixedmlp": 1000,
        "fixedgrad": 1000,
        "fixedada": 1000,
        "fixedlog": 1000}
algs3 = {
        "svm": 1000,
        "grad": 1000,
        "ada": 1000,
        "log": 1000,
        "mlp": 1000,
        "fixedsvm": 1000,
        "fixedmlp": 1000,
        "fixedgrad": 1000,
        "fixedada": 1000,
        "fixedlog": 1000}
algs4 = {
        "grad": 1000,
        "ada": 1000,
        "log": 1000,
        "mlp": 1000,
        "fixedmlp": 1000,
        "fixedgrad": 1000,
        "fixedada": 1000,
        "fixedlog": 1000}
bank = {
    1: algs1,
    2: algs2,
    3: algs3,
    4: algs4
    }

profs = {
        "svm": 0,
        "grad": 0,
        "ada": 0,
        "log": 0,
        "mlp": 0,
        "fixedsvm": 0,
        "fixedmlp": 0,
        "fixedgrad": 0,
        "fixedada": 0,
        "fixedlog": 0}
profit = {
    1: profs,
    2: profs,
    3: profs,
    4: profs
    }

for ii in range(int(sys.argv[-1]), 0):
    record = []
    result = []
    coeffs = []
    player1 = calendar[ii].winner
    player2 = calendar[ii].loser
    for match in calendar[:ii]:
        winner = (match, match.winner)
        loser = (match, match.loser)
        if match.winner == player1:
            record.append([match.lrank - match.wrank, match.welo - match.lelo, match.welosur - match.lelosur, surf_into_num[match.surface], \
                           wins_percent(*winner) - wins_percent(*loser), wins_per_surface(*winner) - wins_per_surface(*loser), \
                           av_first_serve(*winner) - av_first_serve(*loser), av_first_serve_surface(*winner) - av_first_serve_surface(*loser), \
                           av_second_serve(*winner) - av_second_serve(*loser), av_second_serve_surface(*winner) - av_second_serve_surface(*loser), \
                           av_first_return(*winner) - av_first_return(*loser), av_first_return_surface(*winner) - av_first_return_surface(*loser), \
                           av_second_return(*winner) - av_second_return(*loser), av_second_return_surface(*winner) - av_second_return_surface(*loser), \
                           av_aces(*winner) - av_aces(*loser), av_aces_surface(*winner) - av_aces_surface(*loser), \
                           av_dfs(*winner) - av_dfs(*loser), av_dfs_surface(*winner) - av_dfs_surface(*loser), \
                           av_bps(*winner) - av_bps(*loser), av_bps_surface(*winner) - av_bps_surface(*loser)])
            result.append(1)
            coeffs.append([match.cfw, match. cfl])
        elif match.loser == player1:
            record.append([match.wrank - match.lrank, match.lelo-match.welo, match.lelosur-match.welosur, surf_into_num[match.surface], \
                           wins_percent(*loser) - wins_percent(*winner), wins_per_surface(*loser) - wins_per_surface(*winner), \
                           av_first_serve(*loser) - av_first_serve(*winner), av_first_serve_surface(*loser) - av_first_serve_surface(*winner), \
                           av_second_serve(*loser) - av_second_serve(*winner), av_second_serve_surface(*loser) - av_second_serve_surface(*winner), \
                           av_first_return(*loser) - av_first_return(*winner), av_first_return_surface(*loser) - av_first_return_surface(*winner), \
                           av_second_return(*loser) - av_second_return(*winner), av_second_return_surface(*loser) - av_second_return_surface(*winner), \
                           av_aces(*loser) - av_aces(*winner), av_aces_surface(*loser) - av_aces_surface(*winner), \
                           av_dfs(*loser) - av_dfs(*winner), av_dfs_surface(*loser) - av_dfs_surface(*winner), \
                           av_bps(*loser) - av_bps(*winner), av_bps_surface(*loser) - av_bps_surface(*winner)])
            result.append(2)
            coeffs.append([match.cfl, match. cfw])
        if match.loser == player2:
            record.append([match.lrank - match.wrank, match.welo-match.lelo, match.welosur-match.lelosur, surf_into_num[match.surface], \
                           wins_percent(*winner) - wins_percent(*loser), wins_per_surface(*winner) - wins_per_surface(*loser), \
                           av_first_serve(*winner) - av_first_serve(*loser), av_first_serve_surface(*winner) - av_first_serve_surface(*loser), \
                           av_second_serve(*winner) - av_second_serve(*loser), av_second_serve_surface(*winner) - av_second_serve_surface(*loser), \
                           av_first_return(*winner) - av_first_return(*loser), av_first_return_surface(*winner) - av_first_return_surface(*loser), \
                           av_second_return(*winner) - av_second_return(*loser), av_second_return_surface(*winner) - av_second_return_surface(*loser), \
                           av_aces(*winner) - av_aces(*loser), av_aces_surface(*winner) - av_aces_surface(*loser), \
                           av_dfs(*winner) - av_dfs(*loser), av_dfs_surface(*winner) - av_dfs_surface(*loser), \
                           av_bps(*winner) - av_bps(*loser), av_bps_surface(*winner) - av_bps_surface(*loser)])
            result.append(1)
            coeffs.append([match.cfw, match. cfl])
        elif match.winner == player2:
            record.append([match.wrank - match.lrank, match.lelo-match.welo, match.lelosur-match.welosur, surf_into_num[match.surface], \
                           wins_percent(*loser) - wins_percent(*winner), wins_per_surface(*loser) - wins_per_surface(*winner), \
                           av_first_serve(*loser) - av_first_serve(*winner), av_first_serve_surface(*loser) - av_first_serve_surface(*winner), \
                           av_second_serve(*loser) - av_second_serve(*winner), av_second_serve_surface(*loser) - av_second_serve_surface(*winner), \
                           av_first_return(*loser) - av_first_return(*winner), av_first_return_surface(*loser) - av_first_return_surface(*winner), \
                           av_second_return(*loser) - av_second_return(*winner), av_second_return_surface(*loser) - av_second_return_surface(*winner), \
                           av_aces(*loser) - av_aces(*winner), av_aces_surface(*loser) - av_aces_surface(*winner), \
                           av_dfs(*loser) - av_dfs(*winner), av_dfs_surface(*loser) - av_dfs_surface(*winner), \
                           av_bps(*loser) - av_bps(*winner), av_bps_surface(*loser) - av_bps_surface(*winner)])
            result.append(2)
            coeffs.append([match.cfl, match. cfw])



    def roi_1(predictions, q=None, coeffs_test=None):
        roi1=0
        for index in np.where(predictions==y_test)[0].tolist():
            if q==None:
                index_in_record = X.tolist().index(X_test[index].tolist())
                if y_test[index]==1 and coeffs[index_in_record][0]>=1.5:
                    roi1 += coeffs[index_in_record][0]-1
                    #print("+1")
                    #print(coeffs[index_in_record])
                    #print(record[index_in_record])
                elif y_test[index]==2 and coeffs[index_in_record][1]>=1.5:
                    roi1 += coeffs[index_in_record][1]-1
                    #print("+2")
                    #print(coeffs[index_in_record])
                    #print(record[index_in_record])
            else:
                if coeffs_test[0]>=1.5:
                    return round(2*q*(coeffs_test[0]-1))/2
                else:
                    return 0
        for index in np.where(predictions!=y_test)[0].tolist():
            if q==None:
                index_in_record = X.tolist().index(X_test[index].tolist())
                if y_test[index]==1 and coeffs[index_in_record][1]>=1.5:
                    roi1 -= 1
                elif y_test[index]==2 and coeffs[index_in_record][0]>=1.5:
                    roi1 -= 1
                #print(y_test[index])
                #print(record[index_in_record])
            else:
                if coeffs_test[1]>=1.5:
                    return round((-2)*q)/2
                else:
                    return 0
        return roi1

    def roi_2(predictions, q=None, coeffs_test=None):
        roi2=0
        for index in np.where(predictions==y_test)[0].tolist():
            if q==None:
                index_in_record = X.tolist().index(X_test[index].tolist())
                if y_test[index]==1:
                    roi2 += coeffs[index_in_record][0]-1
                    #print("+1")
                    #print(coeffs[index_in_record])
                    #print(record[index_in_record])
                elif y_test[index]==2:
                    roi2 += coeffs[index_in_record][1]-1
                    #print("+2")
                    #print(coeffs[index_in_record])
                    #print(record[index_in_record])
            else:
                return round(2*q*(coeffs_test[0]-1))/2
        for index in np.where(predictions!=y_test)[0].tolist():
            if q==None:
                roi2 -= 1
                #print(y_test[index])
                #print(record[index_in_record])
            else:
                return round((-2)*q)/2
        return roi2

    def roi_3(predictions, q=None, coeffs_test=None):
        roi3=0
        for index in np.where(predictions==y_test)[0].tolist():
            if q==None:
                index_in_record = X.tolist().index(X_test[index].tolist())
                if y_test[index]==1 and coeffs[index_in_record][0]>=1.5:
                    bet = 0.15 - 0.05/(coeffs[index_in_record][0]-1)
                    roi3 += bet*(coeffs[index_in_record][0]-1)
                    #print("+1")
                    #print(coeffs[index_in_record])
                    #print(record[index_in_record])
                elif y_test[index]==2 and coeffs[index_in_record][1]>=1.5:
                    bet = 0.15 - 0.05/(coeffs[index_in_record][1]-1)
                    roi3 += bet*(coeffs[index_in_record][1]-1)
                    #print("+2")
                    #print(coeffs[index_in_record])
                    #print(record[index_in_record])
            else:
                if coeffs_test[0]>=1.5:
                    return round(20*q*(0.15 - 0.05/(coeffs_test[0]-1))*(coeffs_test[0]-1))/2
                else:
                    return 0
        for index in np.where(predictions!=y_test)[0].tolist():
            if q==None:
                index_in_record = X.tolist().index(X_test[index].tolist())
                if y_test[index]==1 and coeffs[index_in_record][1]>=1.5:
                    bet = 0.15 - 0.05/(coeffs[index_in_record][1]-1)
                    roi3 -= bet
                elif y_test[index]==2 and coeffs[index_in_record][0]>=1.5:
                    bet = 0.15 - 0.05/(coeffs[index_in_record][0]-1)
                    roi3 -= bet
                #print(y_test[index])
                #print(record[index_in_record])
            else:
                if coeffs_test[1]>=1.5:
                    return round((-20)*q*(0.15 - 0.05/(coeffs_test[1]-1)))/2
                else:
                    return 0
        return roi3

    def roi_4(predictions, probas, q=None, coeffs_test=None):
        for index in np.where(predictions==y_test)[0].tolist():
            if probas[0][0]>1/coeffs_test[0]:
                return round(2*q*(probas[0][0]*coeffs_test[0]-1))/2
            else:
                return 0
        for index in np.where(predictions!=y_test)[0].tolist():
            if probas[0][1]>1/coeffs_test[1]:
                return round((-2)*q*(probas[0][1]*coeffs_test[1]-1)/(coeffs_test[1]-1))/2
            else:
                return 0
                

    roi_funcs = [None, roi_1, roi_2, roi_3, roi_4]


    class BestCV:

        def __init__(self, params):
            global X,Y, X_test,y_test, X_train,y_train, roi_funcs
            self.X = X
            self.Y = Y
            self.X_test = X_test
            self.y_test = y_test
            self.X_train = X_train
            self.y_train = y_train
            self.roi_funcs = roi_funcs

            self.params = params
            self.best_params_ = None
            
        def search(self):    
            rois = []
            keys = self.params.keys()
            if 'C' in keys and 'gamma' in keys:
                for i in range(len(self.params['C'])):
                    for j in range(len(self.params['gamma'])):
                        clf = SVC(random_state=7, C=self.params['C'][i], gamma=self.params['gamma'][j])
                        clf.fit(self.X_train, self.y_train)
                        predictions = []
                        for match in X_test:
                            predictions.append(clf.predict([match])[0])
                        predictions = np.array(predictions)
                                
                        roi=self.roi_funcs[1](predictions)+self.roi_funcs[3](predictions)
                                    
                        rois.append((roi, self.params['C'][i], self.params['gamma'][j]))
                max_tuple = max(rois, key=lambda x: x[0])
                self.best_params_ = {'C': max_tuple[1], 'gamma': max_tuple[2]}
                return SVC(random_state=7, C=max_tuple[1], gamma=max_tuple[2])
            elif 'learning_rate' in keys and 'n_estimators' in keys and 'min_samples_split' in keys:
                for i in range(len(self.params['learning_rate'])):
                    for j in range(len(self.params['n_estimators'])):
                        for p in range(len(self.params['min_samples_split'])):
                            clf = GBC(random_state=7, learning_rate=self.params['learning_rate'][i], n_estimators=self.params['n_estimators'][j], \
                                      min_samples_split=self.params['min_samples_split'][p])
                            clf.fit(self.X_train, self.y_train)
                            predictions = []
                            for match in X_test:
                                predictions.append(clf.predict([match])[0])
                            predictions = np.array(predictions)
                                    
                            roi=self.roi_funcs[1](predictions)+self.roi_funcs[3](predictions)
                                        
                            rois.append((roi, self.params['learning_rate'][i], self.params['n_estimators'][j], self.params['min_samples_split'][p]))
                max_tuple = max(rois, key=lambda x: x[0])
                self.best_params_ = {'learning_rate': max_tuple[1], 'n_estimators': max_tuple[2], 'min_samples_split': max_tuple[3]}
                return GBC(random_state=7, learning_rate=max_tuple[1], n_estimators=max_tuple[2], min_samples_split=max_tuple[3])
            elif 'learning_rate' in keys and 'n_estimators' in keys and len(keys)==2:
                for i in range(len(self.params['learning_rate'])):
                    for j in range(len(self.params['n_estimators'])):
                        clf = ABC(random_state=7, learning_rate=self.params['learning_rate'][i], n_estimators=self.params['n_estimators'][j])
                        clf.fit(self.X_train, self.y_train)
                        predictions = []
                        for match in X_test:
                            predictions.append(clf.predict([match])[0])
                        predictions = np.array(predictions)
                                
                        roi=self.roi_funcs[1](predictions)+self.roi_funcs[3](predictions)
                                    
                        rois.append((roi, self.params['learning_rate'][i], self.params['n_estimators'][j]))
                max_tuple = max(rois, key=lambda x: x[0])
                self.best_params_ = {'learning_rate': max_tuple[1], 'n_estimators': max_tuple[2]}
                return ABC(random_state=7, learning_rate=max_tuple[1], n_estimators=max_tuple[2])
            elif 'solver' in keys:
                for i in range(len(self.params['C'])):
                    for j in range(len(self.params['solver'])):
                        clf = Log(random_state=7, C=self.params['C'][i], solver=self.params['solver'][j])
                        clf.fit(self.X_train, self.y_train)
                        predictions = []
                        for match in X_test:
                            predictions.append(clf.predict([match])[0])
                        predictions = np.array(predictions)
                                
                        roi=self.roi_funcs[1](predictions)+self.roi_funcs[3](predictions)
                                    
                        rois.append((roi, self.params['C'][i], self.params['solver'][j]))
                max_tuple = max(rois, key=lambda x: x[0])
                self.best_params_ = {'C': max_tuple[1], 'solver': max_tuple[2]}
                return Log(random_state=7, C=max_tuple[1], solver=max_tuple[2])
            else:
                for i in range(len(self.params['hidden_layer_sizes'])):
                    for j in range(len(self.params['activation'])):
                        for p in range(len(self.params['alpha'])):
                            clf = MLP(random_state=7, solver='lbfgs', hidden_layer_sizes=self.params['hidden_layer_sizes'][i], activation=self.params['activation'][j], \
                                      alpha=self.params['alpha'][p])
                            clf.fit(self.X_train, self.y_train)
                            predictions = []
                            for match in X_test:
                                predictions.append(clf.predict([match])[0])
                            predictions = np.array(predictions)
                                    
                            roi=self.roi_funcs[1](predictions)+self.roi_funcs[3](predictions)
                                        
                            rois.append((roi, self.params['hidden_layer_sizes'][i], self.params['activation'][j], self.params['alpha'][p]))
                max_tuple = max(rois, key=lambda x: x[0])
                self.best_params_ = {'hidden_layer_sizes': max_tuple[1], 'activation': max_tuple[2], 'alpha': max_tuple[3]}
                return MLP(random_state=7, solver='lbfgs', hidden_layer_sizes=max_tuple[1], activation=max_tuple[2], alpha=max_tuple[3])
                    
                    
    ##################################

    fr = 1/10

    print('\n', player1, '-', player2, '\n')

    ##################################

    X_pre = np.array(record)
    Y = np.array(result)

    scaler = MinMaxScaler()
    X = scaler.fit_transform(X_pre)
    
    X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.33, random_state=7)

    winner = (calendar[ii], calendar[ii].winner)
    loser = (calendar[ii], calendar[ii].loser)
    pre_predict = [calendar[ii].lrank - calendar[ii].wrank, calendar[ii].welo-calendar[ii].lelo, calendar[ii].welosur-calendar[ii].lelosur, surf_into_num[calendar[ii].surface], \
                   round(wins_percent(*winner) - wins_percent(*loser), 3), round(wins_per_surface(*winner) - wins_per_surface(*loser), 3), \
                   round(av_first_serve(*winner) - av_first_serve(*loser), 3), round(av_first_serve_surface(*winner) - av_first_serve_surface(*loser), 3), \
                   round(av_second_serve(*winner) - av_second_serve(*loser), 3), round(av_second_serve_surface(*winner) - av_second_serve_surface(*loser), 3), \
                   round(av_first_return(*winner) - av_first_return(*loser), 3), round(av_first_return_surface(*winner) - av_first_return_surface(*loser), 3), \
                   round(av_second_return(*winner) - av_second_return(*loser), 3), round(av_second_return_surface(*winner) - av_second_return_surface(*loser), 3), \
                   round(av_aces(*winner) - av_aces(*loser), 3), round(av_aces_surface(*winner) - av_aces_surface(*loser), 3), \
                   round(av_dfs(*winner) - av_dfs(*loser), 3), round(av_dfs_surface(*winner) - av_dfs_surface(*loser), 3), \
                   round(av_bps(*winner) - av_bps(*loser), 3), round(av_bps_surface(*winner) - av_bps_surface(*loser), 3)]
    to_predict = scaler.transform([pre_predict])

    params_svm = {'C': [1e-3, 1e-2, 1e-1, 1, 10, 100, 1000, 10000], 'gamma': [1e-5, 5e-5, 1e-4, 5e-4, 1e-3, 5e-3, 1e-2, 5e-2, 1e-1, 5e-1, 0.85, 1, 10, 100]} 
    model = BestCV(params_svm)

    clf = model.search()
    clf.fit(X, Y)

    X_test = to_predict
    y_test = [1]
    prediction = clf.predict(to_predict)
    coeffs_test = (calendar[ii].cfw, calendar[ii].cfl)

    q1 = fr*bank[1]["svm"]
    q2 = fr*bank[2]["svm"]
    q3 = fr*bank[3]["svm"]

    delta = (None, roi_1(prediction, q1, coeffs_test), roi_2(prediction, q2, coeffs_test), roi_3(prediction, q3, coeffs_test))
    
    print("SVM:")
    print("Best params:", model.best_params_)
    print("To predict:", pre_predict)
    print("Prediction:", prediction)
    print("testing:", "("+str(delta[1]), str(delta[2]), str(delta[3])+")")

    profit[1]["svm"] += delta[1]
    profit[2]["svm"] += delta[2]
    profit[3]["svm"] += delta[3]


    write_profits("SVM", (winner[1], loser[1]), coeffs_test, delta[1:])

    ########################################

    X = np.array(record)
    
    X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.33, random_state=7)

    params_gbc = {'learning_rate': [1e-4, 1e-3, 1e-2, 0.1, 0.2], 'n_estimators': [60, 100, 150, 200], 'min_samples_split': [2]} 
    model = BestCV(params_gbc)

    clf = model.search()
    clf.fit(X, Y)

    X_test = pre_predict
    y_test = [1]
    prediction = clf.predict([pre_predict])
    proba = clf.predict_proba([pre_predict])

    q1 = fr*bank[1]["grad"]
    q2 = fr*bank[2]["grad"]
    q3 = fr*bank[3]["grad"]
    q4 = fr*bank[4]["grad"]

    delta = (None, roi_1(prediction, q1, coeffs_test), roi_2(prediction, q2, coeffs_test), roi_3(prediction, q3, coeffs_test),
             roi_4(prediction, proba, q4, coeffs_test))

    print("\nGradBoost:")
    print("Best params:", model.best_params_)
    print("To predict:", pre_predict)
    print("Prediction:", prediction)
    print("Probabilities:", proba)
    print("testing:", "("+str(delta[1]), str(delta[2]), str(delta[3]), str(delta[4])+")")

    profit[1]["grad"] += delta[1]
    profit[2]["grad"] += delta[2]
    profit[3]["grad"] += delta[3]
    profit[4]["grad"] += delta[4]

    write_profits("GradBoost", (winner[1], loser[1]), coeffs_test, delta[1:])

    ####################################

    X = np.array(record)
    
    X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.33, random_state=7)

    params_abc = {'learning_rate': [1e-3, 1e-2, 1e-1, 0.5, 1, 2], 'n_estimators': [5, 10, 25, 50, 100]} 
    model = BestCV(params_abc)

    clf = model.search()
    clf.fit(X, Y)

    X_test = pre_predict
    y_test = [1]
    prediction = clf.predict([pre_predict])
    proba = clf.predict_proba([pre_predict])

    q1 = fr*bank[1]["ada"]
    q2 = fr*bank[2]["ada"]
    q3 = fr*bank[3]["ada"]
    q4 = fr*bank[4]["ada"]

    delta = (None, roi_1(prediction, q1, coeffs_test), roi_2(prediction, q2, coeffs_test), roi_3(prediction, q3, coeffs_test),
             roi_4(prediction, proba, q4, coeffs_test))

    print("\nAdaBoost:")
    print("Best params:", model.best_params_)
    print("To predict:", pre_predict)
    print("Prediction:", prediction)
    print("Probabilities:", proba)
    print("testing:", "("+str(delta[1]), str(delta[2]), str(delta[3]), str(delta[4])+")")

    profit[1]["ada"] += delta[1]
    profit[2]["ada"] += delta[2]
    profit[3]["ada"] += delta[3]
    profit[4]["ada"] += delta[4]

    write_profits("AdaBoost", (winner[1], loser[1]), coeffs_test, delta[1:])

    ###################################

    X_pre = np.array(record)

    scaler = StandardScaler()
    X = scaler.fit_transform(X_pre)
    
    X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.33, random_state=7)
    
    to_predict = scaler.transform([pre_predict])

    params_log = {'C': [1e-4, 1e-3, 1e-2, 1e-1, 0.5, 1, 10, 100, 1000, 10000], 'solver': ['liblinear', 'newton-cg']} 
    model = BestCV(params_log)

    clf = model.search()
    clf.fit(X, Y)

    X_test = to_predict
    y_test = [1]
    prediction = clf.predict(to_predict)
    proba = clf.predict_proba(to_predict)

    q1 = fr*bank[1]["log"]
    q2 = fr*bank[2]["log"]
    q3 = fr*bank[3]["log"]
    q4 = fr*bank[4]["log"]

    delta = (None, roi_1(prediction, q1, coeffs_test), roi_2(prediction, q2, coeffs_test), roi_3(prediction, q3, coeffs_test),
             roi_4(prediction, proba, q4, coeffs_test))

    print("\nLogReg:")
    print("Best params:", model.best_params_)
    print("To predict:", pre_predict)
    print("Prediction:", prediction)
    print("Probabilities:", proba)
    print("testing:", "("+str(delta[1]), str(delta[2]), str(delta[3]), str(delta[4])+")")

    profit[1]["log"] += delta[1]
    profit[2]["log"] += delta[2]
    profit[3]["log"] += delta[3]
    profit[4]["log"] += delta[4]

    write_profits("LogReg", (winner[1], loser[1]), coeffs_test, delta[1:])

    #################################    

    X_pre = np.array(record)

    scaler = StandardScaler()
    X = scaler.fit_transform(X_pre)
    
    X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.33, random_state=7)

    to_predict = scaler.transform([pre_predict])

    params_mlp = {'hidden_layer_sizes': [(3,), (5,), (8,), (12,), (17,)], \
                  'activation': ['logistic', 'tanh', 'relu'], 'alpha': [1e-5, 1e-4, 1e-3, 1e-2, 1e-1, 0.5, 1, 2, 5, 10],} 
    model = BestCV(params_mlp)

    clf = model.search()
    clf.fit(X, Y)

    X_test = to_predict
    y_test = [1]
    prediction = clf.predict(to_predict)
    proba = clf.predict_proba(to_predict)

    q1 = fr*bank[1]["mlp"]
    q2 = fr*bank[2]["mlp"]
    q3 = fr*bank[3]["mlp"]
    q4 = fr*bank[4]["mlp"]
    
    delta = (None, roi_1(prediction, q1, coeffs_test), roi_2(prediction, q2, coeffs_test), roi_3(prediction, q3, coeffs_test),
             roi_4(prediction, proba, q4, coeffs_test))

    print("\nMLP:")
    print("Best params:", model.best_params_)
    print("To predict:", pre_predict)
    print("Prediction:", prediction)
    print("Probabilities:", proba)
    print("testing:", "("+str(delta[1]), str(delta[2]), str(delta[3]), str(delta[4])+")")

    profit[1]["mlp"] += delta[1]
    profit[2]["mlp"] += delta[2]
    profit[3]["mlp"] += delta[3]
    profit[4]["mlp"] += delta[4]

    write_profits("MLP", (winner[1], loser[1]), coeffs_test, delta[1:])

    ######################################

    X_pre = np.array(record)

    scaler = MinMaxScaler()
    X = scaler.fit_transform(X_pre)
    
    X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.33, random_state=7)

    to_predict = scaler.transform([pre_predict])

    clf = SVC(random_state=7, C=10, gamma=1e-1)
    clf.fit(X, Y)

    X_test = to_predict
    y_test = [1]
    prediction = clf.predict(to_predict)
    coeffs_test = [calendar[ii].cfw, calendar[ii].cfl]

    q1 = fr*bank[1]["fixedsvm"]
    q2 = fr*bank[2]["fixedsvm"]
    q3 = fr*bank[3]["fixedsvm"]

    delta = (None, roi_1(prediction, q1, coeffs_test), roi_2(prediction, q2, coeffs_test), roi_3(prediction, q3, coeffs_test))

    print("\nfixedSVM:")
    print("To predict:", pre_predict)
    print("Prediction:", prediction)
    print("testing:", "("+str(delta[1]), str(delta[2]), str(delta[3])+")")

    profit[1]["fixedsvm"] += delta[1]
    profit[2]["fixedsvm"] += delta[2]
    profit[3]["fixedsvm"] += delta[3]

    write_profits("fixedSVM", (winner[1], loser[1]), coeffs_test, delta[1:])

    ######################################

    X_pre = np.array(record)

    scaler = StandardScaler()
    X = scaler.fit_transform(X_pre)
    
    X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.33, random_state=7)

    to_predict = scaler.transform([pre_predict])

    clf = MLP(random_state=7, solver='lbfgs', hidden_layer_sizes=(12,), activation='logistic', alpha=1e-3)
    clf.fit(X, Y)

    X_test = to_predict
    y_test = [1]
    prediction = clf.predict(to_predict)
    proba = clf.predict_proba(to_predict)
    coeffs_test = [calendar[ii].cfw, calendar[ii].cfl]

    q1 = fr*bank[1]["fixedmlp"]
    q2 = fr*bank[2]["fixedmlp"]
    q3 = fr*bank[3]["fixedmlp"]
    q4 = fr*bank[4]["fixedmlp"]

    delta = (None, roi_1(prediction, q1, coeffs_test), roi_2(prediction, q2, coeffs_test), roi_3(prediction, q3, coeffs_test),
             roi_4(prediction, proba, q4, coeffs_test))

    print("\nfixedMLP:")
    print("To predict:", pre_predict)
    print("Prediction:", prediction)
    print("Probabilities:", proba)
    print("testing:", "("+str(delta[1]), str(delta[2]), str(delta[3]), str(delta[4])+")")

    profit[1]["fixedmlp"] += delta[1]
    profit[2]["fixedmlp"] += delta[2]
    profit[3]["fixedmlp"] += delta[3]
    profit[4]["fixedmlp"] += delta[4]

    write_profits("fixedMLP", (winner[1], loser[1]), coeffs_test, delta[1:])

    ########################################

    X = np.array(record)
    
    X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.33, random_state=7)

    clf = GBC(random_state=7, learning_rate=1e-5, n_estimators=20, min_samples_split=2)
    clf.fit(X, Y)

    X_test = pre_predict
    y_test = [1]
    prediction = clf.predict([pre_predict])
    proba = clf.predict_proba([pre_predict])
    coeffs_test = [calendar[ii].cfw, calendar[ii].cfl]

    q1 = fr*bank[1]["fixedgrad"]
    q2 = fr*bank[2]["fixedgrad"]
    q3 = fr*bank[3]["fixedgrad"]
    q4 = fr*bank[4]["fixedgrad"]

    delta = (None, roi_1(prediction, q1, coeffs_test), roi_2(prediction, q2, coeffs_test), roi_3(prediction, q3, coeffs_test),
             roi_4(prediction, proba, q4, coeffs_test))

    print("\nfixedGrad:")
    print("To predict:", pre_predict)
    print("Prediction:", prediction)
    print("Probabilities:", proba)
    print("testing:", "("+str(delta[1]), str(delta[2]), str(delta[3]), str(delta[4])+")")

    profit[1]["fixedgrad"] += delta[1]
    profit[2]["fixedgrad"] += delta[2]
    profit[3]["fixedgrad"] += delta[3]
    profit[4]["fixedgrad"] += delta[4]

    write_profits("fixedGrad", (winner[1], loser[1]), coeffs_test, delta[1:])

    ###########################################

    X = np.array(record)
    
    X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.33, random_state=7)

    clf = ABC(random_state=7, learning_rate=2, n_estimators=50)
    clf.fit(X, Y)

    X_test = pre_predict
    y_test = [1]
    prediction = clf.predict([pre_predict])
    proba = clf.predict_proba([pre_predict])
    coeffs_test = [calendar[ii].cfw, calendar[ii].cfl]

    q1 = fr*bank[1]["fixedada"]
    q2 = fr*bank[2]["fixedada"]
    q3 = fr*bank[3]["fixedada"]
    q4 = fr*bank[4]["fixedada"]

    delta = (None, roi_1(prediction, q1, coeffs_test), roi_2(prediction, q2, coeffs_test), roi_3(prediction, q3, coeffs_test),
             roi_4(prediction, proba, q4, coeffs_test))

    print("\nfixedAda:")
    print("To predict:", pre_predict)
    print("Prediction:", prediction)
    print("Probabilities:", proba)
    print("testing:", "("+str(delta[1]), str(delta[2]), str(delta[3]), str(delta[4])+")")
    profit[1]["fixedada"] += delta[1]
    profit[2]["fixedada"] += delta[2]
    profit[3]["fixedada"] += delta[3]
    profit[4]["fixedada"] += delta[4]

    write_profits("fixedAda", (winner[1], loser[1]), coeffs_test, delta[1:])

    ########################################

    X_pre = np.array(record)

    scaler = StandardScaler()
    X = scaler.fit_transform(X_pre)
    
    X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.33, random_state=7)

    to_predict = scaler.transform([pre_predict])

    clf = Log(random_state=7, solver='newton-cg', C=10000)
    clf.fit(X, Y)

    X_test = to_predict
    y_test = [1]
    prediction = clf.predict(to_predict)
    proba = clf.predict_proba(to_predict)
    coeffs_test = [calendar[ii].cfw, calendar[ii].cfl]

    q1 = fr*bank[1]["fixedlog"]
    q2 = fr*bank[2]["fixedlog"]
    q3 = fr*bank[3]["fixedlog"]
    q4 = fr*bank[4]["fixedlog"]

    delta = (None, roi_1(prediction, q1, coeffs_test), roi_2(prediction, q2, coeffs_test), roi_3(prediction, q3, coeffs_test),
             roi_4(prediction, proba, q4, coeffs_test))

    print("\nfixedLog:")
    print("To predict:", pre_predict)
    print("Prediction:", prediction)
    print("Probabilities:", proba)
    print("testing:", "("+str(delta[1]), str(delta[2]), str(delta[3]), str(delta[4])+")")

    profit[1]["fixedlog"] += delta[1]
    profit[2]["fixedlog"] += delta[2]
    profit[3]["fixedlog"] += delta[3]
    profit[4]["fixedlog"] += delta[4]

    write_profits("fixedLog", (winner[1], loser[1]), coeffs_test, delta[1:])


'''print("\n", "SVM_profit:", profit[1]["svm"], profit[2]["svm"], profit[3]["svm"])
print("\n", "GradBoost_profit:", profit[1]["grad"], profit[2]["grad"], profit[3]["grad"], profit[4]["grad"])
print("\n", "AdaBoost_profit:", profit[1]["ada"], profit[2]["ada"], profit[3]["ada"], profit[4]["ada"])
print("\n", "LogReg_profit:", profit[1]["log"], profit[2]["log"], profit[3]["log"], profit[4]["log"])
print("\n", "MLP_profit:", profit[1]["mlp"], profit[2]["mlp"], profit[3]["mlp"], profit[4]["mlp"])
print("\n", "fixedSVM_profit:", profit[1]["fixedsvm"], profit[2]["fixedsvm"], profit[3]["fixedsvm"])
print("\n", "MLP_profit:", profit[1]["fixedmlp"], profit[2]["fixedmlp"], profit[3]["fixedmlp"], profit[4]["fixedmlp"])
print("\n", "fixedGrad_profit:", profit[1]["fixedgrad"], profit[2]["fixedgrad"], profit[3]["fixedgrad"], profit[4]["fixedgrad"])
print("\n", "fixedAda_profit:", profit[1]["fixedada"], profit[2]["fixedada"], profit[3]["fixedada"], profit[4]["fixedada"])
print("\n", "fixedLog_profit:", profit[1]["fixedlog"], profit[2]["fixedlog"], profit[3]["fixedlog"], profit[4]["fixedlog"])'''
