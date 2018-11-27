import numpy as np
from sklearn.linear_model import LogisticRegression as Log
from sklearn.metrics import make_scorer
from openpyxl import load_workbook
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from datetime import datetime, timedelta
import matplotlib.pyplot as plt

wb2017 = load_workbook('../Data/2017_simple.xlsx')
ws2017 = wb2017['2017_simple']

wb2018 = load_workbook('../Data/2018_simple.xlsx')
ws2018 = wb2018['2018_simple']

wb2019 = load_workbook('../Data/2019_simple.xlsx')
ws2019 = wb2019['2019_simple']

f = 0.8

def not_NA_rank(rank):
    if rank == 'N/A':
        return 2000
    else:
        return rank

class Match(object):
    def __init__(self, atp=None, location=None, winner=None, loser=None, wrank=None, lrank=None, surface=None, date=None, cfw=None, cfl=None, round_n=None, series=None, \
                 wfsw=None, wssw=None, lfsw=None, lssw=None, wfrw=None, wsrw=None, lfrw=None, lsrw=None, wsrat=None, wrrat=None, lsrat=None, lrrat=None, \
                 waces=None, laces=None, wdf=None, ldf=None, wbp=None, lbp=None, welo=None, lelo=None, bug1=None, bug2=None, welosur=None, lelosur=None):
        self.atp = atp
        self.location = location
        self.winner = winner
        self.loser = loser
        self.wrank = not_NA_rank(wrank)
        self.lrank = not_NA_rank(lrank)
        self.surface = surface
        self.date = date
        self.cfw = cfw
        self.cfl = cfl
        self.round = round_n
        self.series = series
        self.wfsw=wfsw
        self.wssw=wssw
        self.lfsw=lfsw
        self.lssw=lssw
        self.wfrw=wfrw
        self.wsrw=wsrw
        self.lfrw=lfrw
        self.lsrw=lsrw
        self.waces=waces
        self.laces=laces
        self.wdf=wdf
        self.ldf=ldf
        self.wbp=wbp
        self.lbp=lbp
        self.welo=welo
        self.lelo=lelo
        self.welosur=welosur
        self.lelosur=lelosur
    def log(self):
        return dict(atp=self.atp, location=self.location, winner=self.winner, loser=self.loser, wrank=self.wrank, lrank=self.lrank, surface=self.surface, date=self.date, round_n=self.round, series=self.series)

surf_into_num = {'Clay': 0, 'Hard': 0.5, 'Grass': 1}

calendar = []
for row in list(ws2017.rows)[1:]:
    args = [cell.value for cell in row][:36]
    match = Match(*args)
    calendar.append(match)
for row in list(ws2018.rows)[1:]:
    args = [cell.value for cell in row][:36]
    match = Match(*args)
    calendar.append(match)
for row in list(ws2019.rows)[1:]:
    args = [cell.value for cell in row][:36]
    match = Match(*args)
    calendar.append(match)

def wins_per_surface(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    surface = match.surface
    i = 1
    wins = 0
    losses = 0
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match_temp.surface != match.surface:
            i += 1
            continue
        delta = match.date - match_temp.date
        if match.surface == 'Grass':
            t = delta.days/280
        else:
            t = delta.days/7
        if match_temp.winner == player:
            wins += min(f**t, f)
        elif match_temp.loser == player:
            losses += min(f**t, f)
        i += 1
        counter_matches -= 1
    if wins==0:
        return 0
    else:
        return round(wins/(wins+losses), 3)

def wins_percent(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    wins = 0
    losses = 0
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        delta = match.date - match_temp.date
        t = delta.days/7
        if match_temp.winner == player:
            wins += min(f**t, f)
        elif match_temp.loser == player:
            losses += min(f**t, f)
        i += 1
        counter_matches -= 1
    if wins==0:
        return 0
    else:
        return round(wins/(wins+losses), 3)

def av_first_serve(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    first_serve = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        delta = match.date - match_temp.date
        t = delta.days/28
        if match_temp.winner==player:
            first_serve.append((int(match_temp.wfsw[:-1])/100)*min(f**t, f))
        else:
            first_serve.append((int(match_temp.lfsw[:-1])/100)*min(f**t, f))
        i += 1
        counter_matches -= 1
    if len(first_serve)>0:
        return round(sum(first_serve)/len(first_serve), 3)
    else:
        return 0

def av_second_serve(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    second_serve = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        delta = match.date - match_temp.date
        t = delta.days/28
        if match_temp.winner==player:
            second_serve.append((int(match_temp.wssw[:-1])/100)*min(f**t, f))
        else:
            second_serve.append((int(match_temp.lssw[:-1])/100)*min(f**t, f))
        i += 1
        counter_matches -= 1
    if len(second_serve)>0:
        return round(sum(second_serve)/len(second_serve), 3)
    else:
        return 0

def av_first_serve_surface(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    first_serve = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match.surface != match_temp.surface:
            i += 1
            continue
        delta = match.date - match_temp.date
        if match.surface == 'Grass':
            t = 0
        else:
            t = delta.days/28
        if match_temp.winner==player:
            first_serve.append((int(match_temp.wfsw[:-1])/100)*min(f**t, f))
        else:
            first_serve.append((int(match_temp.lfsw[:-1])/100)*min(f**t, f))
        i += 1
        counter_matches -= 1
    if len(first_serve)>0:
        return round(sum(first_serve)/len(first_serve), 3)
    else:
        return 0

def av_second_serve_surface(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    second_serve = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match.surface != match_temp.surface:
            i += 1
            continue
        delta = match.date - match_temp.date
        if match.surface == 'Grass':
            t = 0
        else:
            t = delta.days/28
        if match_temp.winner==player:
            second_serve.append((int(match_temp.wssw[:-1])/100)*min(f**t, f))
        else:
            second_serve.append((int(match_temp.lssw[:-1])/100)*min(f**t, f))
        i += 1
        counter_matches -= 1
    if len(second_serve)>0:
        return round(sum(second_serve)/len(second_serve), 3)
    else:
        return 0

def av_first_return(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    first_serve = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        delta = match.date - match_temp.date
        t = delta.days/28
        if match_temp.winner==player:
            first_serve.append((int(match_temp.wfrw[:-1])/100)*min(f**t, f))
        else:
            first_serve.append((int(match_temp.lfrw[:-1])/100)*min(f**t, f))
        i += 1
        counter_matches -= 1
    if len(first_serve)>0:
        return round(sum(first_serve)/len(first_serve), 3)
    else:
        return 0

def av_second_return(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    second_serve = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        delta = match.date - match_temp.date
        t = delta.days/28
        if match_temp.winner==player:
            second_serve.append((int(match_temp.wsrw[:-1])/100)*min(f**t, f))
        else:
            second_serve.append((int(match_temp.lsrw[:-1])/100)*min(f**t, f))
        i += 1
        counter_matches -= 1
    if len(second_serve)>0:
        return round(sum(second_serve)/len(second_serve), 3)
    else:
        return 0

def av_first_return_surface(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    first_serve = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match.surface != match_temp.surface:
            i += 1
            continue
        delta = match.date - match_temp.date
        if match.surface == 'Grass':
            t = 0
        else:
            t = delta.days/28
        if match_temp.winner==player:
            first_serve.append((int(match_temp.wfrw[:-1])/100)*min(f**t, f))
        else:
            first_serve.append((int(match_temp.lfrw[:-1])/100)*min(f**t, f))
        i += 1
        counter_matches -= 1
    if len(first_serve)>0:
        return round(sum(first_serve)/len(first_serve), 3)
    else:
        return 0

def av_second_return_surface(match, player):
    global f
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    second_serve = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match.surface != match_temp.surface:
            i += 1
            continue
        delta = match.date - match_temp.date
        if match.surface == 'Grass':
            t = 0
        else:
            t = delta.days/28
        if match_temp.winner==player:
            second_serve.append((int(match_temp.wsrw[:-1])/100)*min(f**t, f))
        else:
            second_serve.append((int(match_temp.lsrw[:-1])/100)*min(f**t, f))
        i += 1
        counter_matches -= 1
    if len(second_serve)>0:
        return round(sum(second_serve)/len(second_serve), 3)
    else:
        return 0

def av_aces(match, player):
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    aces = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match_temp.winner==player:
            aces.append(int(match_temp.waces))
        else:
            aces.append(int(match_temp.laces))
        i += 1
        counter_matches -= 1
    if len(aces)>0:
        return round(sum(aces)/len(aces), 3)
    else:
        return 0

def av_aces_surface(match, player):
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    aces = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match.surface != match_temp.surface:
            i += 1
            continue
        if match_temp.winner==player:
            aces.append(float(match_temp.waces))
        else:
            aces.append(float(match_temp.laces))
        i += 1
        counter_matches -= 1
    if len(aces)>0:
        return round(sum(aces)/len(aces), 3)
    else:
        return 0

def av_dfs(match, player):
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    dfs = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match_temp.winner==player:
            dfs.append(float(match_temp.wdf))
        else:
            dfs.append(float(match_temp.ldf))
        i += 1
        counter_matches -= 1
    if len(dfs)>0:
        return round(sum(dfs)/len(dfs), 5)
    else:
        return 0

def av_dfs_surface(match, player):
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    dfs = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match.surface != match_temp.surface:
            i += 1
            continue
        if match_temp.winner==player:
            dfs.append(float(match_temp.wdf))
        else:
            dfs.append(float(match_temp.ldf))
        i += 1
        counter_matches -= 1
    if len(dfs)>0:
        return round(sum(dfs)/len(dfs), 5)
    else:
        return 0

def av_bps(match, player):
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    bps = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match_temp.winner==player:
            bps.append(int(match_temp.wbp[:-1])/100)
        else:
            bps.append(int(match_temp.lbp[:-1])/100)
        i += 1
        counter_matches -= 1
    if len(bps)>0:
        return round(sum(bps)/len(bps), 3)
    else:
        return 0

def av_bps_surface(match, player):
    counter_matches = 40
    match_index = calendar.index(match)
    i = 1
    bps = []
    while counter_matches > 0:
        if match_index - i < 0:
            break
        match_temp = calendar[match_index-i]
        if match_temp.winner != player and match_temp.loser != player:
            i += 1
            continue
        if match.surface != match_temp.surface:
            i += 1
            continue
        if match_temp.winner==player:
            bps.append(int(match_temp.wbp[:-1])/100)
        else:
            bps.append(int(match_temp.lbp[:-1])/100)
        i += 1
        counter_matches -= 1
    if len(bps)>0:
        return round(sum(bps)/len(bps), 3)
    else:
        return 0


bank1 = 1000
bank2 = 1000
bank3 = 1000
bank4 = 1000

profit1 = 0
profit2 = 0
profit3 = 0
profit4 = 0
for ii in range(-1, 0):
    record = []
    result = []
    coeffs = []
    player1 = calendar[ii].winner
    player2 = calendar[ii].loser
    for match in calendar[:-1]:
        winner = (match, match.winner)
        loser = (match, match.loser)
        if match.winner == player1:
            record.append([match.lrank - match.wrank, match.welo - match.lelo, match.welosur - match.lelosur, surf_into_num[match.surface], \
                           wins_percent(*winner) - wins_percent(*loser), wins_per_surface(*winner) - wins_per_surface(*loser), \
                           av_first_serve(*winner) - av_first_serve(*loser), av_first_serve_surface(*winner) - av_first_serve_surface(*loser), \
                           av_second_serve(*winner) - av_second_serve(*loser), av_second_serve_surface(*winner) - av_second_serve_surface(*loser), \
                           av_first_return(*winner) - av_first_return(*loser), av_first_return_surface(*winner) - av_first_return_surface(*loser), \
                           av_second_return(*winner) - av_second_return(*loser), av_second_return_surface(*winner) - av_second_return_surface(*loser), \
                           av_aces(*winner) - av_aces(*loser), av_aces_surface(*winner) - av_aces_surface(*loser), \
                           av_dfs(*winner) - av_dfs(*loser), av_dfs_surface(*winner) - av_dfs_surface(*loser), \
                           av_bps(*winner) - av_bps(*loser), av_bps_surface(*winner) - av_bps_surface(*loser)])
            result.append(1)
            coeffs.append([match.cfw, match. cfl])
        elif match.loser == player1:
            record.append([match.wrank - match.lrank, match.lelo-match.welo, match.lelosur-match.welosur, surf_into_num[match.surface], \
                           wins_percent(*loser) - wins_percent(*winner), wins_per_surface(*loser) - wins_per_surface(*winner), \
                           av_first_serve(*loser) - av_first_serve(*winner), av_first_serve_surface(*loser) - av_first_serve_surface(*winner), \
                           av_second_serve(*loser) - av_second_serve(*winner), av_second_serve_surface(*loser) - av_second_serve_surface(*winner), \
                           av_first_return(*loser) - av_first_return(*winner), av_first_return_surface(*loser) - av_first_return_surface(*winner), \
                           av_second_return(*loser) - av_second_return(*winner), av_second_return_surface(*loser) - av_second_return_surface(*winner), \
                           av_aces(*loser) - av_aces(*winner), av_aces_surface(*loser) - av_aces_surface(*winner), \
                           av_dfs(*loser) - av_dfs(*winner), av_dfs_surface(*loser) - av_dfs_surface(*winner), \
                           av_bps(*loser) - av_bps(*winner), av_bps_surface(*loser) - av_bps_surface(*winner)])
            result.append(2)
            coeffs.append([match.cfl, match. cfw])
        if match.loser == player2:
            record.append([match.lrank - match.wrank, match.welo-match.lelo, match.welosur-match.lelosur, surf_into_num[match.surface], \
                           wins_percent(*winner) - wins_percent(*loser), wins_per_surface(*winner) - wins_per_surface(*loser), \
                           av_first_serve(*winner) - av_first_serve(*loser), av_first_serve_surface(*winner) - av_first_serve_surface(*loser), \
                           av_second_serve(*winner) - av_second_serve(*loser), av_second_serve_surface(*winner) - av_second_serve_surface(*loser), \
                           av_first_return(*winner) - av_first_return(*loser), av_first_return_surface(*winner) - av_first_return_surface(*loser), \
                           av_second_return(*winner) - av_second_return(*loser), av_second_return_surface(*winner) - av_second_return_surface(*loser), \
                           av_aces(*winner) - av_aces(*loser), av_aces_surface(*winner) - av_aces_surface(*loser), \
                           av_dfs(*winner) - av_dfs(*loser), av_dfs_surface(*winner) - av_dfs_surface(*loser), \
                           av_bps(*winner) - av_bps(*loser), av_bps_surface(*winner) - av_bps_surface(*loser)])
            result.append(1)
            coeffs.append([match.cfw, match. cfl])
        elif match.winner == player2:
            record.append([match.wrank - match.lrank, match.lelo-match.welo, match.lelosur-match.welosur, surf_into_num[match.surface], \
                           wins_percent(*loser) - wins_percent(*winner), wins_per_surface(*loser) - wins_per_surface(*winner), \
                           av_first_serve(*loser) - av_first_serve(*winner), av_first_serve_surface(*loser) - av_first_serve_surface(*winner), \
                           av_second_serve(*loser) - av_second_serve(*winner), av_second_serve_surface(*loser) - av_second_serve_surface(*winner), \
                           av_first_return(*loser) - av_first_return(*winner), av_first_return_surface(*loser) - av_first_return_surface(*winner), \
                           av_second_return(*loser) - av_second_return(*winner), av_second_return_surface(*loser) - av_second_return_surface(*winner), \
                           av_aces(*loser) - av_aces(*winner), av_aces_surface(*loser) - av_aces_surface(*winner), \
                           av_dfs(*loser) - av_dfs(*winner), av_dfs_surface(*loser) - av_dfs_surface(*winner), \
                           av_bps(*loser) - av_bps(*winner), av_bps_surface(*loser) - av_bps_surface(*winner)])
            result.append(2)
            coeffs.append([match.cfl, match. cfw])



    def roi_1(predictions, q=None, coeffs_test=None):
        roi1=0
        for index in np.where(predictions==y_test)[0].tolist():
            if q==None:
                index_in_record = X.tolist().index(X_test[index].tolist())
                if y_test[index]==1 and coeffs[index_in_record][0]>=1.5:
                    roi1 += coeffs[index_in_record][0]-1
                    #print("+1")
                    #print(coeffs[index_in_record])
                    #print(record[index_in_record])
                elif y_test[index]==2 and coeffs[index_in_record][1]>=1.5:
                    roi1 += coeffs[index_in_record][1]-1
                    #print("+2")
                    #print(coeffs[index_in_record])
                    #print(record[index_in_record])
            else:
                if coeffs_test[0]>=1.5:
                    return round(2*q*(coeffs_test[0]-1))/2
                else:
                    return 0
        for index in np.where(predictions!=y_test)[0].tolist():
            if q==None:
                index_in_record = X.tolist().index(X_test[index].tolist())
                if y_test[index]==1 and coeffs[index_in_record][1]>=1.5:
                    roi1 -= 1
                elif y_test[index]==2 and coeffs[index_in_record][0]>=1.5:
                    roi1 -= 1
                #print(y_test[index])
                #print(record[index_in_record])
            else:
                if coeffs_test[1]>=1.5:
                    return round((-2)*q)/2
                else:
                    return 0
        return roi1

    def roi_2(predictions, q=None, coeffs_test=None):
        roi2=0
        for index in np.where(predictions==y_test)[0].tolist():
            if q==None:
                index_in_record = X.tolist().index(X_test[index].tolist())
                if y_test[index]==1:
                    roi2 += coeffs[index_in_record][0]-1
                    #print("+1")
                    #print(coeffs[index_in_record])
                    #print(record[index_in_record])
                elif y_test[index]==2:
                    roi2 += coeffs[index_in_record][1]-1
                    #print("+2")
                    #print(coeffs[index_in_record])
                    #print(record[index_in_record])
            else:
                return round(2*q*(coeffs_test[0]-1))/2
        for index in np.where(predictions!=y_test)[0].tolist():
            if q==None:
                roi2 -= 1
                #print(y_test[index])
                #print(record[index_in_record])
            else:
                return round((-2)*q)/2
        return roi2

    def roi_3(predictions, q=None, coeffs_test=None):
        roi3=0
        for index in np.where(predictions==y_test)[0].tolist():
            if q==None:
                index_in_record = X.tolist().index(X_test[index].tolist())
                if y_test[index]==1 and coeffs[index_in_record][0]>=1.5:
                    bet = 0.15 - 0.05/(coeffs[index_in_record][0]-1)
                    roi3 += bet*(coeffs[index_in_record][0]-1)
                    #print("+1")
                    #print(coeffs[index_in_record])
                    #print(record[index_in_record])
                elif y_test[index]==2 and coeffs[index_in_record][1]>=1.5:
                    bet = 0.15 - 0.05/(coeffs[index_in_record][1]-1)
                    roi3 += bet*(coeffs[index_in_record][1]-1)
                    #print("+2")
                    #print(coeffs[index_in_record])
                    #print(record[index_in_record])
            else:
                if coeffs_test[0]>=1.5:
                    return round(20*q*(0.15 - 0.05/(coeffs_test[0]-1))*(coeffs_test[0]-1))/2
                else:
                    return 0
        for index in np.where(predictions!=y_test)[0].tolist():
            if q==None:
                index_in_record = X.tolist().index(X_test[index].tolist())
                if y_test[index]==1 and coeffs[index_in_record][1]>=1.5:
                    bet = 0.15 - 0.05/(coeffs[index_in_record][1]-1)
                    roi3 -= bet
                elif y_test[index]==2 and coeffs[index_in_record][0]>=1.5:
                    bet = 0.15 - 0.05/(coeffs[index_in_record][0]-1)
                    roi3 -= bet
                #print(y_test[index])
                #print(record[index_in_record])
            else:
                if coeffs_test[1]>=1.5:
                    return round((-20)*q*(0.15 - 0.05/(coeffs_test[1]-1)))/2
                else:
                    return 0
        return roi3

    def roi_4(predictions, probas, q=None, coeffs_test=None):
        for index in np.where(predictions==y_test)[0].tolist():
            if probas[0][0]>1/coeffs_test[0]:
                return round(2*q*(probas[0][0]*coeffs_test[0]-1))/2
            else:
                return 0
        for index in np.where(predictions!=y_test)[0].tolist():
            if probas[0][1]>1/coeffs_test[1]:
                return round((-2)*q*(probas[0][1]*coeffs_test[1]-1)/(coeffs_test[1]-1))/2
            else:
                return 0
                

    roi_funcs = [None, roi_1, roi_2, roi_3]


    class BestCV:

        def __init__(self, params):
            global X,Y, X_test,y_test, X_train,y_train, roi_funcs
            self.X = X
            self.Y = Y
            self.X_test = X_test
            self.y_test = y_test
            self.X_train = X_train
            self.y_train = y_train
            self.roi_funcs = roi_funcs

            self.params = params
            self.best_params_ = None
            
        def search(self):    
            rois = []
            for i in range(len(self.params['C'])):
                for j in range(len(self.params['solver'])):
                    clf = Log(random_state=7, C=self.params['C'][i], solver=self.params['solver'][j])
                    clf.fit(self.X_train, self.y_train)
                    predictions = []
                    for match in X_test:
                        predictions.append(clf.predict([match])[0])
                    predictions = np.array(predictions)
                            
                    roi=self.roi_funcs[1](predictions)+self.roi_funcs[3](predictions)
                                
                    rois.append((roi, self.params['C'][i], self.params['solver'][j]))
            max_tuple = max(rois, key=lambda x: x[0])
            self.best_params_ = {'C': max_tuple[1], 'solver': max_tuple[2]}
            return Log(random_state=7, C=max_tuple[1], solver=max_tuple[2])
                    

    X_pre = np.array(record)
    Y = np.array(result)

    scaler = StandardScaler()
    X = scaler.fit_transform(X_pre)
    
    X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.33, random_state=7)

    winner = (calendar[ii], calendar[ii].winner)
    loser = (calendar[ii], calendar[ii].loser)
    pre_predict = [calendar[ii].lrank - calendar[ii].wrank, calendar[ii].welo-calendar[ii].lelo, calendar[ii].welosur-calendar[ii].lelosur, surf_into_num[calendar[ii].surface], \
                   round(wins_percent(*winner) - wins_percent(*loser), 3), round(wins_per_surface(*winner) - wins_per_surface(*loser), 3), \
                   round(av_first_serve(*winner) - av_first_serve(*loser), 3), round(av_first_serve_surface(*winner) - av_first_serve_surface(*loser), 3), \
                   round(av_second_serve(*winner) - av_second_serve(*loser), 3), round(av_second_serve_surface(*winner) - av_second_serve_surface(*loser), 3), \
                   round(av_first_return(*winner) - av_first_return(*loser), 3), round(av_first_return_surface(*winner) - av_first_return_surface(*loser), 3), \
                   round(av_second_return(*winner) - av_second_return(*loser), 3), round(av_second_return_surface(*winner) - av_second_return_surface(*loser), 3), \
                   round(av_aces(*winner) - av_aces(*loser), 3), round(av_aces_surface(*winner) - av_aces_surface(*loser), 3), \
                   round(av_dfs(*winner) - av_dfs(*loser), 5), round(av_dfs_surface(*winner) - av_dfs_surface(*loser), 5), \
                   round(av_bps(*winner) - av_bps(*loser), 3), round(av_bps_surface(*winner) - av_bps_surface(*loser), 3)]
    to_predict = scaler.transform([pre_predict])

    params_log = {'C': [1e-4, 1e-3, 1e-2, 1e-1, 0.5, 1, 10, 100, 1000, 10000], 'solver': ['liblinear', 'newton-cg']} 
    model = BestCV(params_log)

    clf = model.search()
    clf.fit(X, Y)

    #print(round(clf.score(X_test, y_test), 3))

    #predictions = []
    #probas = []
    #for match in X_test:
    #    predictions.append(clf.predict([match])[0])
    #    probas.append(clf.predict_proba([match])[0])
    #predictions = np.array(predictions)
    #probas = np.array(probas)


    #print("roi1: {}%".format(int(round(100*roi_1(predictions), 0))))
    #print("roi2: {}%".format(int(round(100*roi_2(predictions), 0))))
    #print("roi3: {}%".format(int(round(1000*roi_3(predictions), 0))))

    X_test = to_predict
    y_test = [1]
    prediction = clf.predict(to_predict)
    proba = clf.predict_proba(to_predict)
    coeffs_test = [calendar[ii].cfw, calendar[ii].cfl]

    q1 = 0.1*bank1
    q2 = 0.1*bank2
    q3 = 0.1*bank3
    q4 = 0.1*bank4
    
    print("Best params:", model.best_params_)
    print("To predict:", pre_predict)
    print("Prediction:", prediction)
    print("Probabilities:", proba)
    print("testing:", "("+str(roi_1(prediction, q1, coeffs_test)), str(roi_2(prediction, q2, coeffs_test)), str(roi_3(prediction, q3, coeffs_test)), \
          str(roi_4(prediction, proba, q4, coeffs_test))+")")

    profit1 += roi_1(prediction, q1, coeffs_test)
    profit2 += roi_2(prediction, q2, coeffs_test)
    profit3 += roi_3(prediction, q3, coeffs_test)
    profit4 += roi_4(prediction, proba, q4, coeffs_test)

print("\n", profit1, profit2, profit3, profit4)
